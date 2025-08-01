import numpy as np
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from pandas import DataFrame
from pandas import IndexSlice
from pandas.io.excel import read_excel
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
import matplotlib.pyplot as plt

#Объявление переменных и значения по умолчанию
#filename=''
f_tables='output_tables.txt'
f_out = 'output_vibr.txt'

J = C = l = a = phi = pg = Pg = []
n_res_strong_1 = n_res_strong_2 = n_res_strong_3 = []
n_res_main_1 = n_res_main_2 = n_res_main_3 = []

N=72
K=24
L=10

#все параметры (кроме расхода топлива) в СИ

#количество моторных масс
M = 6
#модуль упругости
G = 8.3*10**10
#количество цилиндров
I = 6
#диаметр коренной шейки
dk = 0.1
#момент инерции маховика демпфера
J_M = 0.01
Jp = w_s = w_01 = w_nom = 0
#радиус кривошипа
R = 0.1
#диаметр цилиндра
D = 0.1
#параметр подобия
lambd = 1.0
#номинальная частота оборотов
n_nom = 10000
#масса возвратно-поступательно движущихся частей
Mvp = 0.7
#минимальная частота оборотов
n_min = 6000
#максимальная частота оборотов
n_max = 12000
#тактность
tau = 4
#объем цилиндра
V = 0.0005
#номинальная мощность
Ne_nom = 100000.0
#коэффициент демпфирования
eps = 5.0
#номинальный удельный расход топлива (в г/(кВт*ч))
ge_nom=2.0

#код для вывода ascii-таблиц

def print_pretty_table(data, cell_sep=' | ', header_separator=True):
    rows = len(data)
    cols = len(data[0])

    col_width = []
    for col in range(cols):
        columns = [data[row][col] for row in range(rows)]
        col_width.append(len(max(columns, key=len)))

    separator = "-+-".join('-' * n for n in col_width)

    for i, row in enumerate(range(rows)):
        if i == 1 and header_separator:
            print(separator)

        result = []
        for col in range(cols):
            item = data[row][col].rjust(col_width[col])
            result.append(item)

        print(cell_sep.join(result))
    print('\n')

def write_pretty_table_to_txt(data, filename, cell_sep=' | ', header_separator=True):
    rows = len(data)
    cols = len(data[0])

    col_width = []
    for col in range(cols):
        columns = [data[row][col] for row in range(rows)]
        col_width.append(len(max(columns, key=len)))

    separator = "-+-".join('-' * n for n in col_width)

    with open(filename, 'a') as file:
        for i, row in enumerate(range(rows)):
            if i == 1 and header_separator:
                file.write(separator + '\n')

            result = []
            for col in range(cols):
                item = data[row][col].rjust(col_width[col])
                result.append(item)

            file.write(cell_sep.join(result) + '\n')
        file.write('\n')

def write_to_file(data, filename):
    with open(filename, 'a') as file:
        file.write(data)        

#проверка вводимого значения на целое число
def is_valid_number(newval):
    return newval == '' or newval.isnumeric()

#проверка вводимого значения на вещественное число
def is_valid_float(newval):
    try:
        float(newval)
        f = True
    except:
        f = False
    return newval == '' or f

#функция формирует массив a[i]
def find_a(w):
    global a
    a=np.ones(M)
    for i in range(1, M):
      aJ=a*J
      a[i]=a[i-1]-w**2*np.sum(aJ[0:i])/C[i-1]
    return a

#минимизируемая ошибка
def min_to_find(w):
    Ja=J*find_a(w)
    return abs(np.sum(Ja)/(np.max(abs(Ja))))

#поиск собственной частоты
def find_w_s():
    global w_s
    w_s = w_01
    while (w_s>0 and min_to_find(w_s)>0.05):
        w_s = w_s-w_01*0.01

    if min_to_find(w_s)>0.05:
          while (w_s<10*w_01 and min_to_find(w_s)>0.05):
              w_s = w_s+w_01*0.01
    

def print_Jp():
    #Jp_print=Jp*10**8
    Jp_label.config(text='Полярный момент инерции Jp = ' + ("%.8f" %Jp)+' м\u2074')
    Jp_label.grid(column=4, row=0, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('Полярный момент инерции Jp = ' + ("%.8f" %Jp)+' м^4\n', f_out)
        

def print_Ji():
    global wb
    str1=['i']+list(map(str, range(1,M+1)))
    str2=['J_i,кг*м^2']+list(map(str, J.tolist()))
    data=[str1,str2]

    print('Моменты инерции масс в кг*м^2\n\n')
    print_pretty_table(data)
    
    if textfile_on.get():
        
        write_to_file('Моменты инерции масс в кг*м^2\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)
        
    if excelfile_on.get():
        str1=['i']+list(map(str, range(1,M+1)))
        str2=['J_i, кг\u00B7м\u00B2']+list(map(str, J.tolist()))
        data=[str1,str2]
        
        df = DataFrame(data[1:], columns=data[0])
        wb = Workbook()
        ws = wb.active
        ws.title = 'Моменты инерции Ji'

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')

def print_JlC():
    global wb
    
    str1=['i,i+1']
    for i in range(1,M):
        nums=str(i)+','+str(i+1)
        str1.append(nums)
    C_print=C.tolist()
    l_print=l.tolist()
    C_print=[str(round(num)) for num in C]
    l_print=["{:.3f}".format(num) for num in l]
  
    str2=['C_i,i+1, Н*м/рад']+C_print
    str3=['l_i,i+1, м']+l_print
    data=[str1,str2,str3]

    print('Жесткости и длины промежутков\n\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Жесткости и длины промежутков\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=['i,i+1']
        for i in range(1,M):
            nums=str(i)+','+str(i+1)
            str1.append(nums)
        C_print=C.tolist()
        l_print=l.tolist()
        C_print=[str(round(num)) for num in C]
        l_print=["{:.3f}".format(num) for num in l]
      
        str2=['C_i,i+1, Н\u00B7м/рад']+C_print
        str3=['l_i,i+1, м']+l_print
        data=[str1,str2,str3]
        
        ws = wb.create_sheet(title='Длины и жесткости')
        data=[str1,str2,str3]
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')

def print_w_s():
    w_s_label.config(text='Собственная частота w_s = ' + ("%.0f" %w_s)+' рад/с')
    w_s_label.grid(column=4, row=1, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('Собственная частота w_s = ' + ("%.8f" %w_s)+' рад/с\n', f_out)

def print_a():
    global wb
    str1=['i']+list(map(str, range(1,M+1)))
    str2=['a_i']+list(map(str, (np.round(a,3)).tolist()))
    data=[str1, str2]
    
    print('Коэффициенты a_i\n\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Коэффициенты a_i\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():     
        ws = wb.create_sheet(title='Коэффициенты a_i')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')

def print_n_res():
    global wb
    data=[['K_m', 'n_res, об/м']]
    
    data.append(['K_m_strong_1']+[str(round(n_res_strong_1))])
    data.append(['K_m_main_1']+[str(round(n_res_main_1))])

    data.append(['K_m_strong_2']+[str(round(n_res_strong_2))])
    data.append(['K_m_main_2']+[str(round(n_res_main_2))])

    data.append(['K_m_strong_3']+[str(round(n_res_strong_3))])
    data.append(['K_m_main_3']+[str(round(n_res_main_3))])

    print('Сильные и главные гармоники\n\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Сильные и главные гармоники\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)
    if excelfile_on.get():     
        ws = wb.create_sheet(title='Сильные и главные гармоники')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_M_rot_g_mean():
    M_rot_g_mean_label.config(text='Средний крутящий газовый момент = ' + ("%.2f" %M_rot_g_mean)+' Н\u00B7м')
    M_rot_g_mean_label.grid(column=4, row=2, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('Средний крутящий газовый момент = ' + ("%.2f" %M_rot_g_mean)+' Н*м\n', f_out)

def print_M_rot_j_mean():
    M_rot_j_mean_label.config(text='Средний инерционный газовый момент = ' + ("%.2f" %M_rot_j_mean)+' Н\u00B7м')
    M_rot_j_mean_label.grid(column=4, row=3, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('Средний инерционный газовый момент = ' + ("%.2f" %M_rot_j_mean)+' Н*м\n', f_out)

def print_Fourier_table():
    global wb
    data=[['k', 'М_кр_газ_k, Н*м', 'phi_газ_k', 'М_кр_инерц_k, Н*м', 'phi_инерц_k']]
    for k in range(K):
        data.append([str(k+1), str(round(M_rot_g_sqr[k],2)), str(round(phi_g[k],2)), str(round(M_rot_j_sqr[k],2)), str(round(phi_j[k],2))])
    print('Моменты гармоник\n\n')
    print_pretty_table(data)
    
    if textfile_on.get():
        write_to_file('Моменты гармоник\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        ws = wb.create_sheet(title='Моменты гармоник')
        
        data[0]=['k', 'М_кр_газ_k, Н\u00B7м', 'phi_газ_k', 'М_кр_инерц_k, Н\u00B7м', 'phi_инерц_k']
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')
    

def print_M_rot_n_g():
    global wb
    str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(M_rot_n_g_strong_1).astype(int).astype(str).tolist()]
    str3=[['K_main_1']+np.round(M_rot_n_g_main_1).astype(int).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(M_rot_n_g_strong_2).astype(int).astype(str).tolist()]
    str5=[['K_main_2']+np.round(M_rot_n_g_main_2).astype(int).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(M_rot_n_g_strong_3).astype(int).astype(str).tolist()]
    str7=[['K_main_3']+np.round(M_rot_n_g_main_3).astype(int).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('Моменты газовых гармоник в Н*м\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Моменты газовых гармоник в Н*м\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['Первая сильная, Н\u00B7м']+np.round(M_rot_n_g_strong_1).astype(int).astype(str).tolist()]
        str3=[['Первая главная, Н\u00B7м']+np.round(M_rot_n_g_main_1).astype(int).astype(str).tolist()]
        str4=[['Вторая сильная, Н\u00B7м']+np.round(M_rot_n_g_strong_2).astype(int).astype(str).tolist()]
        str5=[['Вторая главная, Н\u00B7м']+np.round(M_rot_n_g_main_2).astype(int).astype(str).tolist()]
        str6=[['Третья сильная, Н\u00B7м']+np.round(M_rot_n_g_strong_3).astype(int).astype(str).tolist()]
        str7=[['Третья главная, Н\u00B7м']+np.round(M_rot_n_g_main_3).astype(int).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7
        ws = wb.create_sheet(title='Моменты газовых гармоник')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_M_rot_n_j():
    global wb
    str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(M_rot_n_j_strong_1).astype(int).astype(str).tolist()]
    str3=[['K_main_1']+np.round(M_rot_n_j_main_1).astype(int).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(M_rot_n_j_strong_2).astype(int).astype(str).tolist()]
    str5=[['K_main_2']+np.round(M_rot_n_j_main_2).astype(int).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(M_rot_n_j_strong_3).astype(int).astype(str).tolist()]
    str7=[['K_main_3']+np.round(M_rot_n_j_main_3).astype(int).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('Моменты инерционных гармоник в Н*м\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Моменты инерционных гармоник в Н*м\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['Первая сильная, Н\u00B7м']+np.round(M_rot_n_j_strong_1).astype(int).astype(str).tolist()]
        str3=[['Первая главная, Н\u00B7м']+np.round(M_rot_n_j_main_1).astype(int).astype(str).tolist()]
        str4=[['Вторая сильная, Н\u00B7м']+np.round(M_rot_n_j_strong_2).astype(int).astype(str).tolist()]
        str5=[['Вторая главная, Н\u00B7м']+np.round(M_rot_n_j_main_2).astype(int).astype(str).tolist()]
        str6=[['Третья сильная, Н\u00B7м']+np.round(M_rot_n_j_strong_3).astype(int).astype(str).tolist()]
        str7=[['Третья главная, Н\u00B7м']+np.round(M_rot_n_j_main_3).astype(int).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7
        ws = wb.create_sheet(title='Моменты инерционных гармоник')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_Ai(i):
    str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(A_strong_1[i],5).astype(float).astype(str).tolist()]
    str3=[['K_main_1']+np.round(A_main_1[i],5).astype(float).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(A_strong_2[i],5).astype(float).astype(str).tolist()]
    str5=[['K_main_2']+np.round(A_main_2[i],5).astype(float).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(A_strong_3[i],5).astype(float).astype(str).tolist()]
    str7=[['K_main_3']+np.round(A_main_3[i],5).astype(float).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('Амплитуда колебаний для i =',str(i+1),'в м\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Амплитуда колебаний для i = '+str(i+1)+' в м\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

def print_excel_Ai():
    data=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    
    for i in range(M):
        str1=[['i = '+str(i+1)]+['']*(L)]
        str2=[['Первая сильная, м']+np.round(A_strong_1[i],5).astype(float).astype(str).tolist()]
        str3=[['Первая главная м']+np.round(A_main_1[i],5).astype(float).astype(str).tolist()]
        str4=[['Вторая сильная м']+np.round(A_strong_2[i],5).astype(float).astype(str).tolist()]
        str5=[['Вторая главная м']+np.round(A_main_2[i],5).astype(float).astype(str).tolist()]
        str6=[['Третья сильная м']+np.round(A_strong_3[i],5).astype(float).astype(str).tolist()]
        str7=[['Третья главная м']+np.round(A_main_3[i],5).astype(float).astype(str).tolist()]
        str8=[['']*(L+1)]
        data+=str1+str2+str3+str4+str5+str6+str7+str8

    ws = wb.create_sheet(title='Амплитуды колебаний')
        
    df = DataFrame(data[1:], columns=data[0])

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
    #wb.save('output.xlsx')
    
def print_M_upr(i):
    str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(M_upr_strong_1[i],1).astype(float).astype(str).tolist()]
    str3=[['K_main_1']+np.round(M_upr_main_1[i],1).astype(float).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(M_upr_strong_2[i],1).astype(float).astype(str).tolist()]
    str5=[['K_main_2']+np.round(M_upr_main_2[i],1).astype(float).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(M_upr_strong_3[i],1).astype(float).astype(str).tolist()]
    str7=[['K_main_3']+np.round(M_upr_main_3[i],1).astype(float).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('Упругие моменты для i =',str(i+1),'в Н*м\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Упругие моменты для i = '+str(i+1)+' в Н*м\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

def print_excel_M_upr():
    data=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    
    for i in range(M-1):
        str1=[['i, i+1 = '+str(i+1)+', '+str(i+2)]+['']*(L)]
        str2=[['Первая сильная, м']+np.round(M_upr_strong_1[i],5).astype(float).astype(str).tolist()]
        str3=[['Первая главная м']+np.round(M_upr_main_1[i],5).astype(float).astype(str).tolist()]
        str4=[['Вторая сильная м']+np.round(M_upr_strong_2[i],5).astype(float).astype(str).tolist()]
        str5=[['Вторая главная м']+np.round(M_upr_main_2[i],5).astype(float).astype(str).tolist()]
        str6=[['Третья сильная м']+np.round(M_upr_strong_3[i],5).astype(float).astype(str).tolist()]
        str7=[['Третья главная м']+np.round(M_upr_main_3[i],5).astype(float).astype(str).tolist()]
        str8=[['']*(L+1)]
        data+=str1+str2+str3+str4+str5+str6+str7+str8

    ws = wb.create_sheet(title='Упругие напряжения')
        
    df = DataFrame(data[1:], columns=data[0])

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
    #wb.save('output.xlsx')

def print_T():
    T_label.config(text='Максимальное напряжение T = ' + ("%.0f" %T)+' Па')
    T_label.grid(column=4, row=4, sticky=tk.W, padx=10)
    if textfile_on.get():
        write_to_file('Максимальное напряжение T = ' + ("%.0f" %T)+' Па\n', f_out)

def print_N():
    str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(N_strong_1).astype(int).astype(str).tolist()]
    str3=[['K_main_1']+np.round(N_main_1).astype(int).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(N_strong_2).astype(int).astype(str).tolist()]
    str5=[['K_main_2']+np.round(N_main_2).astype(int).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(N_strong_3).astype(int).astype(str).tolist()]
    str7=[['K_main_3']+np.round(N_main_3).astype(int).astype(str).tolist()]
    str8=[['Сумма']+np.round(N_sum).astype(int).astype(str).tolist()]
    
    data=str1+str2+str3+str4+str5+str6+str7+str8

    print('Потери мощности в Вт\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Потери мощности в Вт\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)
    if excelfile_on.get():
        str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['Первая сильная, Вт']+np.round(N_strong_1).astype(int).astype(str).tolist()]
        str3=[['Первая главная, Вт']+np.round(N_main_1).astype(int).astype(str).tolist()]
        str4=[['Вторая сильная, Вт']+np.round(N_strong_2).astype(int).astype(str).tolist()]
        str5=[['Вторая главная, Вт']+np.round(N_main_2).astype(int).astype(str).tolist()]
        str6=[['Третья сильная, Вт']+np.round(N_strong_3).astype(int).astype(str).tolist()]
        str7=[['Третья главная, Вт']+np.round(N_main_3).astype(int).astype(str).tolist()]
        str8=[['Суммарная, Вт']+np.round(N_sum).astype(int).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7+str8
        ws = wb.create_sheet(title='Потери мощности')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_dg():
    str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(dg_strong_1,3).astype(float).astype(str).tolist()]
    str3=[['K_main_1']+np.round(dg_main_1,3).astype(float).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(dg_strong_2,3).astype(float).astype(str).tolist()]
    str5=[['K_main_2']+np.round(dg_main_2,3).astype(float).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(dg_strong_3,3).astype(float).astype(str).tolist()]
    str7=[['K_main_3']+np.round(dg_main_3,3).astype(float).astype(str).tolist()]
    str8=[['Сумма']+np.round(dg_sum,3).astype(float).astype(str).tolist()]
    
    data=str1+str2+str3+str4+str5+str6+str7+str8

    print('Перерасход топлива в г/(кВт*ч)\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('Перерасход топлива в г/(кВт*ч)\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=[['n, об/м']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['Первая сильная, г/(кВт\u00B7ч)']+np.round(dg_strong_1, 3).astype(float).astype(str).tolist()]
        str3=[['Первая главная, г/(кВт\u00B7ч)']+np.round(dg_main_1, 3).astype(float).astype(str).tolist()]
        str4=[['Вторая сильная, г/(кВт\u00B7ч)']+np.round(dg_strong_2, 3).astype(float).astype(str).tolist()]
        str5=[['Вторая главная, г/(кВт\u00B7ч)']+np.round(dg_main_2, 3).astype(float).astype(str).tolist()]
        str6=[['Третья сильная, г/(кВт\u00B7ч)']+np.round(dg_strong_3, 3).astype(float).astype(str).tolist()]
        str7=[['Третья главная, г/(кВт\u00B7ч)']+np.round(dg_main_3, 3).astype(float).astype(str).tolist()]
        str8=[['Суммарный, г/(кВт\u00B7ч)']+np.round(dg_sum, 3).astype(float).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7+str8
        ws = wb.create_sheet(title='Перерасход топлива')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        wb.save('output.xlsx')
        
#Запуск расчета
def calculate():
    global N, K, L
    global M, G, I, dk, Jp, J_M, R, D, lambd, n_nom, Mvp, n_min, n_max, V, tau, Ne_nom, eps, g_e_nom
    global J, C, l, n_diap
    global w_01, w_nom
    global j, beta, Pj, Pg, M_rot_g, M_rot_j, Ag, Bg, M_rot_g_sqr, phi_g, M_rot_g_mean
    global M_rot_n_g_strong_1, M_rot_n_g_strong_2, M_rot_n_g_strong_3
    global M_rot_n_g_main_1, M_rot_n_g_main_2, M_rot_n_g_main_3
    global M_rot_n_j_strong_1, M_rot_n_j_strong_2, M_rot_n_j_strong_3
    global M_rot_n_j_main_1, M_rot_n_j_main_2, M_rot_n_j_main_3
    global Aj, Bj, M_rot_j_sqr, phi_j, M_rot_j_mean

    global A1_strong_1, A1_strong_2, A1_strong_3
    global A1_main_1, A1_main_2, A1_main_3
    global A_strong_1, A_strong_2, A_strong_3
    global A_main_1, A_main_2, A_main_3
    
    #гармоники
    global K_m_main_1, K_m_main_2, K_m_main_3
    global K_m_strong_1, K_m_strong_2, K_m_strong_3
    #резонансные обороты
    global n_res_strong_1, n_res_strong_2, n_res_strong_3
    global n_res_main_1, n_res_main_2, n_res_main_3
    
    global phi, pg
    global a_coef, b_coef, c_coef
    global M_upr_strong_1, M_upr_strong_2, M_upr_strong_3
    global M_upr_main_1, M_upr_main_2, M_upr_main_3
    global T
        
    global N_strong_1, N_strong_2, N_strong_3
    global N_main_1, N_main_2, N_main_3
    global N_sum

    global dg_strong_1, dg_strong_2, dg_strong_3
    global dg_main_1, dg_main_2, dg_main_3
    global dg_sum
    
    #задаваемые параметры в СИ
    M = int(entry_M.get())
    G = 10**9*float(entry_G.get())
    dk = float(entry_dk.get())
    J_M = float(entry_JM.get())
    R = float(entry_R.get())
    D = float(entry_D.get())
    lambd = float(entry_lambd.get())
    n_nom = int(entry_n_nom.get())
    Mvp = float(entry_Mvp.get())
    n_min = int(entry_n_min.get())
    n_max = int(entry_n_max.get())
    V = float(entry_V.get())/1000
    tau = int(entry_tau.get())
    Ne_nom = float(entry_Ne_nom.get())*1000
    eps = float(entry_eps.get())
    ge_nom = float(entry_ge_nom.get())
    
    input_file = filedialog.askopenfilename()
    if input_file=='':
        return 0
    open(input_file, "r")

    #open(f_out, "w")
    #open(f_tables, "w")

    if textfile_on.get():
        open(f_tables, 'w')
        open(f_out, 'w')

    a_coef=1
    b_coef=1
    c_coef=-1
    if engine_box.get()=='Дизельный':
        a_coef=0.87
        b_coef=1.13
        c_coef=-1
    
    J = np.array(read_excel(input_file, sheet_name=0, header=None).iloc[1:, 1])
    C = np.array(read_excel(input_file, sheet_name=1, header=None).iloc[1:, 1])
    print('\n')
    #print(J)
    #print(C)
    
    #1. Эквивалентная расчетная схема
    
    Jp = np.pi*dk**4/32
    #здесь будут длины l_i,i+1 между массами
    l = np.zeros(M-1); 
    #вычисляем длины
    l = G*Jp/C

    #3. Собственная частота эквивалентной схемы
    print(J[0:M-1])
    J_sum = np.sum(J[0:M-1])
    l_sum = np.sum(l[0:M-2])+l[M-2]
    C_sum = G*Jp/l_sum

    #приближенное значение частоты
    w_01 = np.sqrt(C_sum*(J_sum+J_M)/(J_sum*J_M))
    find_w_s()

    #4. Главные и сильные гармоники
    
    K_m_main_1 = I*1; K_m_main_2 = I*2; K_m_main_3 = I*3; 
    K_m_strong_1 = int(0.5*I*1); K_m_strong_2 = int(0.5*I*2); K_m_strong_3 = int(0.5*I*3);

    #5. Резонансные обороты

    n_res_strong_1 = 30*w_s/(np.pi*K_m_strong_1)
    n_res_main_1 = 30*w_s/(np.pi*K_m_main_1)

    n_res_strong_2 = 30*w_s/(np.pi*K_m_strong_2)
    n_res_main_2 = 30*w_s/(np.pi*K_m_main_2)

    n_res_strong_3 = 30*w_s/(np.pi*K_m_strong_3)
    n_res_main_3 = 30*w_s/(np.pi*K_m_main_3)

    #6. Гармонический анализ

    #переводим в радианы!
    phi = (np.pi/180)*(np.array(read_excel(input_file, sheet_name=2, header=None).iloc[1:, 0]))
    
    pg = np.array(read_excel(input_file, sheet_name=2, header=None).iloc[1:, 1])
    
    w_nom = 2*np.pi*n_nom/60
    phi=np.array(phi.tolist())
    
    j = R*w_nom**2*(np.cos(phi)+lambd*np.cos(2*phi))
    beta = np.arcsin(lambd*np.sin(phi))
    
    Pj = -Mvp*j
    Pg = 0.25*np.pi*D**2*(pg-100000)
    
    M_rot_g = Pg*np.sin(phi+beta)*R/np.cos(beta)
    M_rot_j = Pj*np.sin(phi+beta)*R/np.cos(beta)

    #разложение в ряд (газовая составляющая)
    Ag = np.zeros(K)
    Bg = np.zeros(K)
    for k in range(K):
        for i in range(N):
            Ag[k]+= 2/N*M_rot_g[i]*np.cos((k+1)*2*np.pi*(i+1)/N)
            Bg[k]+= 2/N*M_rot_g[i]*np.sin((k+1)*2*np.pi*(i+1)/N)

    M_rot_g_sqr = np.sqrt(Ag**2+Bg**2)
    phi_g = np.arctan(Ag/Bg)

    #средний крутящий газовый момент
    M_rot_g_mean = (1/N)*sum(M_rot_g)

    #разложение в ряд (инерционная составляющая)
    Aj = np.zeros(K)
    Bj = np.zeros(K)
    for k in range(K):
        for i in range(N):
            Aj[k]+= 2/N*M_rot_j[i]*np.cos((k+1)*2*np.pi*(i+1)/N)
            Bj[k]+= 2/N*M_rot_j[i]*np.sin((k+1)*2*np.pi*(i+1)/N)

    M_rot_j_sqr = np.sqrt(Aj**2+Bj**2)
    phi_j = np.arctan(Aj/Bj)

    #средний крутящий инерционный момент
    M_rot_j_mean = 1/N*sum(M_rot_j)

    #Среднее давление в рабочем диапазоне
    n_diap=np.linspace(n_min, n_max, L)

    P_e=(30*tau/(V*I))*(Ne_nom/n_nom)*(a_coef+b_coef*n_diap/n_nom+c_coef*(n_diap/n_nom)**2)
    P_e_nom=(30*tau/(V*I))*(Ne_nom/n_nom)*(a_coef+b_coef+c_coef)

    #вычисляем гармоники газовых составляющих
    M_rot_n_g_strong_1=M_rot_g_sqr[K_m_strong_1-1]*P_e/P_e_nom
    M_rot_n_g_strong_2=M_rot_g_sqr[K_m_strong_2-1]*P_e/P_e_nom
    M_rot_n_g_strong_3=M_rot_g_sqr[K_m_strong_3-1]*P_e/P_e_nom

    M_rot_n_g_main_1=M_rot_g_sqr[K_m_main_1-1]*P_e/P_e_nom
    M_rot_n_g_main_2=M_rot_g_sqr[K_m_main_2-1]*P_e/P_e_nom
    M_rot_n_g_main_3=M_rot_g_sqr[K_m_main_3-1]*P_e/P_e_nom

    #вычисляем гармоники инерционных составляющих
    M_rot_n_j_strong_1=M_rot_j_sqr[round(K_m_strong_1)-1]*(n_diap/n_nom)**2
    M_rot_n_j_strong_2=M_rot_j_sqr[round(K_m_strong_2)-1]*(n_diap/n_nom)**2
    M_rot_n_j_strong_3=M_rot_j_sqr[round(K_m_strong_3)-1]*(n_diap/n_nom)**2

    M_rot_n_j_main_1=M_rot_j_sqr[round(K_m_main_1)-1]*(n_diap/n_nom)**2
    M_rot_n_j_main_2=M_rot_j_sqr[round(K_m_main_2)-1]*(n_diap/n_nom)**2
    M_rot_n_j_main_3=M_rot_j_sqr[round(K_m_main_3)-1]*(n_diap/n_nom)**2

    #7. Амплитуда колебаний моторных масс

    A1_strong_1=(M_rot_n_j_strong_1+M_rot_n_g_strong_1)*sum(a)/(eps*K_m_strong_1*(2*np.pi*n_res_strong_1/60)*sum(a**2))
    A1_main_1=(M_rot_n_j_main_1+M_rot_n_g_main_1)*sum(a)/(eps*K_m_main_1*(2*np.pi*n_res_main_1/60)*sum(a**2))

    A1_strong_2=(M_rot_n_j_strong_2+M_rot_n_g_strong_2)*sum(a)/(eps*K_m_strong_2*(2*np.pi*n_res_strong_2/60)*sum(a**2))
    A1_main_2=(M_rot_n_j_main_2+M_rot_n_g_main_2)*sum(a)/(eps*K_m_main_2*(2*np.pi*n_res_main_2/60)*sum(a**2))

    A1_strong_3=(M_rot_n_j_strong_3+M_rot_n_g_strong_3)*sum(a)/(eps*K_m_strong_3*(2*np.pi*n_res_strong_3/60)*sum(a**2))
    A1_main_3=(M_rot_n_j_main_3+M_rot_n_g_main_3)*sum(a)/(eps*K_m_main_3*(2*np.pi*n_res_main_3/60)*sum(a**2))

    A_strong_1=np.zeros((M, L))
    A_main_1=np.zeros((M, L))
    A_strong_2=np.zeros((M, L))
    A_main_2=np.zeros((M, L))
    A_strong_3=np.zeros((M, L))
    A_main_3=np.zeros((M, L))

    for i in range(M):
        A_strong_1[i]=a[i]*A1_strong_1
        A_strong_2[i]=a[i]*A1_strong_2
        A_strong_3[i]=a[i]*A1_strong_3
        A_main_1[i]=a[i]*A1_main_1
        A_main_2[i]=a[i]*A1_main_2
        A_main_3[i]=a[i]*A1_main_3

    #упругие моменты
    M_upr_strong_1=np.zeros((M-1, L))
    M_upr_main_1=np.zeros((M-1, L))
    M_upr_strong_2=np.zeros((M-1, L))
    M_upr_main_2=np.zeros((M-1, L))
    M_upr_strong_3=np.zeros((M-1, L))
    M_upr_main_3=np.zeros((M-1, L))

    for i in range(M-1):
        for j in range(L):
            M_upr_strong_1[i][j]=C[i]*(A_strong_1[i+1][j]-A_strong_1[i][j])
            M_upr_main_1[i][j]=C[i]*(A_main_1[i+1][j]-A_main_1[i][j])
            M_upr_strong_2[i][j]=C[i]*(A_strong_2[i+1][j]-A_strong_2[i][j])
            M_upr_main_2[i][j]=C[i]*(A_main_2[i+1][j]-A_main_2[i][j])
            M_upr_strong_3[i][j]=C[i]*(A_strong_3[i+1][j]-A_strong_3[i][j])
            M_upr_main_3[i][j]=C[i]*(A_main_3[i+1][j]-A_main_3[i][j])

    #касательные напряжения
    T_upr_strong_1 = np.amax(np.abs(M_upr_strong_1))*16/np.pi/dk**3
    T_upr_main_1 = np.amax(np.abs(M_upr_main_1))*16/np.pi/dk**3
    T_upr_strong_2 = np.amax(np.abs(M_upr_strong_2))*16/np.pi/dk**3
    T_upr_main_2 = np.amax(np.abs(M_upr_main_2))*16/np.pi/dk**3
    T_upr_strong_3 = np.amax(np.abs(M_upr_strong_3))*16/np.pi/dk**3
    T_upr_main_3 = np.amax(np.abs(M_upr_main_3))*16/np.pi/dk**3

    #максимальное напряжение
    T=np.amax(np.array([T_upr_strong_1, T_upr_main_1, T_upr_strong_2, T_upr_main_2, T_upr_strong_3, T_upr_main_3]))

    #9. Потери мощности и перерасход топлива

    N_strong_1=(M_rot_g_sqr[round(K_m_strong_1)-1]+M_rot_j_sqr[round(K_m_strong_1)-1])*sum(a)*A1_strong_1*K_m_strong_1*w_s/2
    N_main_1=(M_rot_g_sqr[round(K_m_main_1)-1]+M_rot_j_sqr[round(K_m_main_1)-1])*sum(a)*A1_main_1*K_m_main_1*w_s/2

    N_strong_2=(M_rot_g_sqr[round(K_m_strong_2)-1]+M_rot_j_sqr[round(K_m_strong_2)-1])*sum(a)*A1_strong_2*K_m_strong_2*w_s/2
    N_main_2=(M_rot_g_sqr[round(K_m_main_2)-1]+M_rot_j_sqr[round(K_m_main_2)-1])*sum(a)*A1_main_2*K_m_main_2*w_s/2

    N_strong_3=(M_rot_g_sqr[round(K_m_strong_3)-1]+M_rot_j_sqr[round(K_m_strong_3)-1])*sum(a)*A1_strong_3*K_m_strong_3*w_s/2
    N_main_3=(M_rot_g_sqr[round(K_m_main_3)-1]+M_rot_j_sqr[round(K_m_main_3)-1])*sum(a)*A1_main_1*K_m_main_3*w_s/2

    N_sum=N_strong_1+N_main_1+N_strong_2+N_main_2+N_strong_3+N_main_3

    #перерасход

    Ne=Ne_nom*(a_coef+b_coef*n_diap/n_nom+c_coef*(n_diap/n_nom)**2)
    
    dg_strong_1=ge_nom*N_strong_1/Ne
    dg_main_1=ge_nom*N_main_1/Ne

    dg_strong_2=ge_nom*N_strong_2/Ne
    dg_main_2=ge_nom*N_main_2/Ne

    dg_strong_3=ge_nom*N_strong_3/Ne
    dg_main_3=ge_nom*N_main_3/Ne
    
    dg_sum=dg_strong_1+dg_main_1+dg_strong_2+dg_main_2+dg_strong_3+dg_main_3
       
    print_Jp()
    print_Ji()
    print_JlC()
    print_w_s()
    print_a()
    print_n_res()
    print_M_rot_g_mean()
    print_M_rot_j_mean()
    print_Fourier_table()
    print_M_rot_n_g()
    print_M_rot_n_j()
    for i in range(M):
        print_Ai(i)
    if excelfile_on.get():
        print_excel_Ai()
        print_excel_M_upr()
    for i in range(M-1):
        print_M_upr(i)
    print_T()
    print_N()
    print_dg()

    graphics()

    #plt.figure(num='A1_main_1')
    #plt.plot(n_diap, A1_main_1)
    #plt.title('A1_main_1')
    #plt.show()

#Опция построения графиков
def graphics():
    global graph_label, harm_label, masses_label, upr_label
    global graph_box, harm_box, masses_box, upr_box
    global graph, harm, masses, inter

    harm=['Первая сильная']
    harm.append('Первая главная')
    harm.append('Вторая сильная')
    harm.append('Вторая главная')
    harm.append('Третья сильная')
    harm.append('Третья главная')

    graph_label = ttk.Label(root, text='Построение графика:')
    graph_button = ttk.Button(root, text='Построить', command=make_graph)

    harm_label = ttk.Label(root, text='Гармоника:')
    harm_box = ttk.Combobox(root, width=40, values=harm)
    harm_box.set('Первая сильная')

    masses=[]
    for i in range(M):
        masses.append(str(i+1))
    masses_label = ttk.Label(root, text='Номер массы:')
    masses_box = ttk.Combobox(root, width=40, values=masses)
    masses_box.set('1')

    inter=[]
    for i in range(M-1):
        inter.append(str(i+1))
    upr_label = ttk.Label(root, text='Интервал:')
    upr_box = ttk.Combobox(root, width=40, values=inter)
    upr_box.set('1')

    graph = ['Гармоники газового момента']
    graph.append('Гармоники инерционного момента')
    graph.append('Моменты главных газовых гармоник')
    graph.append('Моменты главных инерционных гармоник')
    graph.append('Амплитуды моторных масс')
    graph.append('Дополнительные упругие напряжения')
    graph.append('Потери мощности')
    graph.append('Перерасход топлива')
    graph.append('Суммарные потери мощности')
    graph.append('Суммарный перерасход топлива')
    
    graph_box = ttk.Combobox(root, width=40, values=graph)
    graph_box.bind("<<ComboboxSelected>>", graphboxchanged)
    
    graph_label.grid(column=4, row=7, sticky=tk.W, padx=10)
    
    graph_box.set('Гармоники газового момента')
    graph_box.grid(column=4, row=8, sticky=tk.W, padx=10, pady=5)
    graph_box.state(["readonly"])
    harm_box.state(["readonly"])
    masses_box.state(["readonly"])

    graph_button.grid(column=4, row=15, columnspan=1, pady=10)
        
def graphboxchanged(event):
    global graph_label, harm_label, masses_label, upr_label
    global graph_box, harm_box, masses_box, upr_box
    global graph, harm, masses, inter
    
    list1=['Гармоники газового момента', 'Гармоники инерционного момента',
           'Суммарные потери мощности', 'Суммарный перерасход топлива']
    list2=['Моменты главных газовых гармоник', 'Моменты главных инерционных гармоник',
           'Потери мощности', 'Перерасход топлива']
    list3=['Амплитуды моторных масс']
    list4=['Дополнительные упругие напряжения']
    
    if graph_box.get() in list1:
        harm_label.grid_forget()
        harm_box.grid_forget()
        masses_label.grid_forget()
        masses_box.grid_forget()
        upr_label.grid_forget()
        upr_box.grid_forget()
    if graph_box.get() in list2:
        harm_label.grid(column=4, row=9, sticky=tk.W, padx=10)
        harm_box.grid(column=4, row=10, sticky=tk.W, padx=10, pady=5)
        masses_label.grid_forget()
        masses_box.grid_forget()
        upr_label.grid_forget()
        upr_box.grid_forget()
    if graph_box.get() in list3:
        upr_label.grid_forget()
        upr_box.grid_forget()
        harm_label.grid(column=4, row=9, sticky=tk.W, padx=10)
        harm_box.grid(column=4, row=10, sticky=tk.W, padx=10, pady=5)
        masses_label.grid(column=4, row=11, sticky=tk.W, padx=10)
        masses_box.grid(column=4, row=12, sticky=tk.W, padx=10, pady=5)
    if graph_box.get() in list4:
        masses_label.grid_forget()
        masses_box.grid_forget()
        harm_label.grid(column=4, row=9, sticky=tk.W, padx=10)
        harm_box.grid(column=4, row=10, sticky=tk.W, padx=10, pady=5)
        upr_label.grid(column=4, row=11, sticky=tk.W, padx=10)
        upr_box.grid(column=4, row=12, sticky=tk.W, padx=10, pady=5)
        
def make_graph():
    global graph_label, harm_label, masses_label, upr_label, dN_label
    global graph_box, harm_box, masses_box, upr_box
    global graph, harm, masses, inter

    if graph_box.get()=='Гармоники газового момента':
        x = list(range(1, K+1))
        plt.figure(num='Гармоники газового момента')
        plt.grid(True)
        plt.plot(x, M_rot_g_sqr)
        plt.title('Гармоники газового момента')
        plt.xlabel('Номер гармоники')
        plt.ylabel('Момент, Н\u00B7м')
        plt.show()
        
    if graph_box.get()=='Гармоники инерционного момента':
        x = list(range(1, 25))
        plt.figure(num='Гармоники инерционного момента')
        plt.grid(True)
        plt.plot(x, M_rot_j_sqr)
        plt.title('Гармоники инерционного момента')
        plt.xlabel('Номер гармоники')
        plt.ylabel('Момент, Н\u00B7м')
        plt.show()

    dict_Mng = {'Первая сильная':M_rot_n_g_strong_1,
        'Первая главная':M_rot_n_g_main_1,
        'Вторая сильная':M_rot_n_g_strong_2,
        'Вторая главная':M_rot_n_g_main_2,
        'Третья сильная':M_rot_n_g_strong_3,
        'Третья главная':M_rot_n_g_main_3}
    
    if graph_box.get()=='Моменты главных газовых гармоник':
        x = n_diap
        plt.figure(num='Моменты главных газовых гармоник. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Mng[harm_box.get()])
        plt.title('Моменты главных газовых гармоник')
        plt.xlabel('n, об/м')
        plt.ylabel('Момент, Н\u00B7м')
        plt.show()

    dict_Mnj = {'Первая сильная':M_rot_n_j_strong_1,
        'Первая главная':M_rot_n_j_main_1,
        'Вторая сильная':M_rot_n_j_strong_2,
        'Вторая главная':M_rot_n_j_main_2,
        'Третья сильная':M_rot_n_j_strong_3,
        'Третья главная':M_rot_n_j_main_3}
    
    if graph_box.get()=='Моменты главных инерционных гармоник':
        x = n_diap
        plt.figure(num='Моменты главных инерционных гармоник. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Mnj[harm_box.get()])
        plt.title('Моменты главных инерционных гармоник')
        plt.xlabel('n, об/м')
        plt.ylabel('Момент, Н\u00B7м')
        plt.show()

    i=int(masses_box.get())-1
    dict_Ai = {'Первая сильная':A_strong_1[i],
        'Первая главная':A_main_1[i],
        'Вторая сильная':A_strong_2[i],
        'Вторая главная':A_main_2[i],
        'Третья сильная':A_strong_3[i],
        'Третья главная':A_main_3[i]
    }
    
    if graph_box.get()=='Амплитуды моторных масс':
        x = n_diap
        plt.figure(num='Амплитуды моторных масс. Масса ' + str(masses_box.get())+'. '+harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Ai[harm_box.get()])
        plt.title('Амплитуды моторных масс. Масса ' + str(masses_box.get())+'. '+harm_box.get()+'.')
        plt.xlabel('n, об/м')
        plt.ylabel('Амплитуда, м')
        plt.show()

    i=int(upr_box.get())-1
    
    dict_Mi = {'Первая сильная':M_upr_strong_1[i],
        'Первая главная':M_upr_main_1[i],
        'Вторая сильная':M_upr_strong_2[i],
        'Вторая главная':M_upr_main_2[i],
        'Третья сильная':M_upr_strong_3[i],
        'Третья главная':M_upr_main_3[i]
    }
    
    if graph_box.get()=='Дополнительные упругие напряжения':
        x = n_diap
        plt.figure(num='Дополнительные упругие напряжения. Интервал ' + str(upr_box.get())+'. '+harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Mi[harm_box.get()])
        plt.title('Дополнительные упругие напряжения. Интервал ' + str(upr_box.get())+'. '+harm_box.get()+'.')
        plt.xlabel('n, об/м')
        plt.ylabel('Напряжение, Па')
        plt.show()

    dict_dN = {'Первая сильная':N_strong_1,
        'Первая главная':N_main_1,
        'Вторая сильная':N_strong_2,
        'Вторая главная':N_main_2,
        'Третья сильная':N_strong_3,
        'Третья главная':N_main_3}

    if graph_box.get()=='Потери мощности':
        x = n_diap
        plt.figure(num='Потери мощности. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_dN[harm_box.get()])
        plt.title('Потери мощности. ' + harm_box.get()+'.')
        plt.xlabel('n, об/м')
        plt.ylabel('Потеря мощности, Вт')
        plt.show()

    dict_dg = {'Первая сильная':dg_strong_1,
        'Первая главная':dg_main_1,
        'Вторая сильная':dg_strong_2,
        'Вторая главная':dg_main_2,
        'Третья сильная':dg_strong_3,
        'Третья главная':dg_main_3}

    if graph_box.get()=='Перерасход топлива':
        x = n_diap
        plt.figure(num='Перерасход топлива. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_dg[harm_box.get()])
        plt.title('Перерасход топлива. ' + harm_box.get()+'.')
        plt.xlabel('n, об/м')
        plt.ylabel('Перерасход, г/(кВт\u00B7ч)')
        plt.show()

    if graph_box.get()=='Суммарные потери мощности':
        x = n_diap
        plt.figure(num='Суммарные потери мощности.')
        plt.grid(True)
        plt.plot(x, N_sum)
        plt.title('Суммарные потери мощности.')
        plt.xlabel('n, об/м')
        plt.ylabel('Потеря мощности, Вт')
        plt.show()

    if graph_box.get()=='Суммарный перерасход топлива':
        x = n_diap
        plt.figure(num='Суммарный перерасход топлива.')
        plt.grid(True)
        plt.plot(x, dg_sum)
        plt.title('Суммарный перерасход топлива.')
        plt.xlabel('n, об/м')
        plt.ylabel('Перерасход, г/(кВт\u00B7ч)')
        plt.show()
    
        
# Создание главного окна
root = tk.Tk()
root. resizable(False, False)

check_number = (root.register(is_valid_number), "%P")
check_float = (root.register(is_valid_float), "%P")

# Создание текстовых меток и полей ввода
label_M = ttk.Label(root, text='Количество моторных масс M =')
label_M.grid(column=0, row=0, sticky=tk.W, padx=10, pady=5)
entry_M = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_M.insert(0, M)
entry_M.grid(column=1, row=0, padx=10, pady=5)

label_G = ttk.Label(root, text='Модуль сдвига G (ГПа) =')
label_G.grid(column=2, row=0, sticky=tk.W)
entry_G = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_G.insert(0, G/10**9)
entry_G.grid(column=3, row=0, padx=10)

label_I = ttk.Label(root, text='Количество цилиндров I =')
label_I.grid(column=0, row=1, sticky=tk.W, padx=10, pady=5)
entry_I = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_I.insert(0, I)
entry_I.grid(column=1, row=1, padx=10, pady=5)

label_dk = ttk.Label(root, text='Диаметр коренной шейки d (м) =')
label_dk.grid(column=2, row=1, sticky=tk.W, pady=5)
entry_dk = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_dk.insert(0, dk)
entry_dk.grid(column=3, row=1, padx=10, pady=5)

label_JM = ttk.Label(root, text='Момент инерции маховика демпфера J_M (кг\u00B7м\u00B2) =')
label_JM.grid(column=0, row=2, sticky=tk.W, padx=10, pady=5)
entry_JM = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_JM.insert(0, J_M)
entry_JM.grid(column=1, row=2, padx=10, pady=5)

label_R = ttk.Label(root, text='Радиус кривошипа R (м) =')
label_R.grid(column=2, row=2, sticky=tk.W, pady=5)
entry_R = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_R.insert(0, R)
entry_R.grid(column=3, row=2, padx=10, pady=5)

label_D = ttk.Label(root, text='Диаметр цилиндра D (м) =')
label_D.grid(column=0, row=3, sticky=tk.W, padx=10, pady=5)
entry_D = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_D.insert(0, D)
entry_D.grid(column=1, row=3, padx=10, pady=5)

label_lambd = ttk.Label(root, text='Параметр подобия \u03BB =')
label_lambd.grid(column=2, row=3, sticky=tk.W, pady=5)
entry_lambd = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_lambd.insert(0, lambd)
entry_lambd.grid(column=3, row=3, padx=10, pady=5)

label_Mvp = ttk.Label(root, text='Масса движущихся частей Mvp (кг) =')
label_Mvp.grid(column=0, row=4, sticky=tk.W, padx=10, pady=5)
entry_Mvp = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_Mvp.insert(0, Mvp)
entry_Mvp.grid(column=1, row=4, padx=10, pady=5)

label_n_nom = ttk.Label(root, text='Номинальная частота оборотов (об/м) =')
label_n_nom.grid(column=2, row=4, sticky=tk.W, pady=5)
entry_n_nom = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_n_nom.insert(3, n_nom)
entry_n_nom.grid(column=3, row=4, padx=10, pady=5)

label_n_min = ttk.Label(root, text='Минимальная частота оборотов (об/м) =')
label_n_min.grid(column=0, row=5, sticky=tk.W, padx=10, pady=5)
entry_n_min = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_n_min.insert(0, n_min)
entry_n_min.grid(column=1, row=5, padx=10, pady=5)

label_n_max = ttk.Label(root, text='Максимальная частота оборотов (об/м) =')
label_n_max.grid(column=2, row=5, sticky=tk.W, pady=5)
entry_n_max = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_n_max.insert(3, n_max)
entry_n_max.grid(column=3, row=5, padx=10, pady=5)

label_V = ttk.Label(root, text='Объем цилиндра V (л) =')
label_V.grid(column=0, row=6, sticky=tk.W, padx=10, pady=5)
entry_V = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_V.insert(0, V*1000)
entry_V.grid(column=1, row=6, padx=10, pady=5)

label_tau = ttk.Label(root, text='Тактность \u03C4 =')
label_tau.grid(column=2, row=6, sticky=tk.W, pady=5)
entry_tau = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_tau.insert(3, tau)
entry_tau.grid(column=3, row=6, padx=10, pady=5)

label_Ne_nom = ttk.Label(root, text='Номинальная мощность Ne_nom (кВт) =')
label_Ne_nom.grid(column=0, row=7, sticky=tk.W, padx=10, pady=5)
entry_Ne_nom = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_Ne_nom.insert(0, Ne_nom/1000)
entry_Ne_nom.grid(column=1, row=7, padx=10, pady=5)

label_eps = ttk.Label(root, text='Коэффициент демпфирования \u03B5 =')
label_eps.grid(column=2, row=7, sticky=tk.W, pady=5)
entry_eps = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_eps.insert(0, eps)
entry_eps.grid(column=3, row=7, padx=10, pady=5)

label_ge_nom = ttk.Label(root, text='Номинальный расход топлива в г/(кВт\u00B7ч) =')
label_ge_nom.grid(column=0, row=8, sticky=tk.W, padx=10, pady=5)
entry_ge_nom = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_ge_nom.insert(0, ge_nom)
entry_ge_nom.grid(column=1, row=8, padx=10, pady=5)

label_ge_nom = ttk.Label(root, text='Тип двигателя:')
label_ge_nom.grid(column=2, row=8, sticky=tk.W, pady=5)
engine=['Искровой', 'Дизельный']
engine_box = ttk.Combobox(root, values=engine)
engine_box.grid(column=3, row=8, sticky=tk.W, padx=10, pady=5)
engine_box.set('Искровой')
engine_box.state(["readonly"])

label_ge_nom = ttk.Label(root, text='Компоновочная схема:')
label_ge_nom.grid(column=2, row=9, sticky=tk.W, pady=5)
scheme=['R', 'V', 'O']
scheme_box = ttk.Combobox(root, values=scheme)
scheme_box.grid(column=3, row=9, sticky=tk.W, padx=10, pady=5)
scheme_box.set('R')
scheme_box.state(["readonly"])

label_file = ttk.Label(root, text='Выберите типы файлов для вывода:')
label_file.grid(column=0, row=15, sticky=tk.W, padx=10, pady=5)

textfile_on = tk.BooleanVar()
text_checkbutton = ttk.Checkbutton(root, text='Текстовый файл', variable=textfile_on)
#text_checkbutton.pack()
text_checkbutton.grid(column=1, row=15, padx=10, pady=5)

excelfile_on = tk.BooleanVar()
excel_checkbutton = ttk.Checkbutton(root, text='Файл Excel', variable=excelfile_on)
excel_checkbutton.grid(column=2, row=15, padx=10, pady=5)


# Создание кнопки для расчета суммы
calculate_button = ttk.Button(root, text='Рассчитать', command=calculate)
calculate_button.grid(column=3, row=15, columnspan=1, pady=10)

# Создание текстового поля для отображения результата
Jp_label = ttk.Label(root)
w_s_label = ttk.Label(root)
M_rot_g_mean_label = ttk.Label(root)
M_rot_j_mean_label = ttk.Label(root)
T_label = ttk.Label(root)


# Запуск главного цикла обработки событий
root.mainloop()
