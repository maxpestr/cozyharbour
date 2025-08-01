import os
import sys
from datetime import datetime
import numpy as np
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from pandas import DataFrame
from pandas import IndexSlice
from pandas.io.excel import read_excel
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import NamedStyle
from openpyxl.styles import Font
import matplotlib.pyplot as plt
from tkinter import Menu
from tkinter.filedialog import asksaveasfilename, askopenfilename
from tkinter import Menu
from tkinter.filedialog import asksaveasfilename, askopenfilename


#–û–±—ä—è–≤–ª–µ–Ω–∏–µ –ø–µ—Ä–µ–º–µ–Ω–Ω—ã—Ö –∏ –∑–Ω–∞—á–µ–Ω–∏—è –ø–æ —É–º–æ–ª—á–∞–Ω–∏—é
#filename=''
f_tables='output_tables.txt'
f_out = 'output_vibr.txt'
params_file_default_name = 'params.xlsx'

J = C = l = a = phi = pg = Pg = []
n_res_strong_1 = n_res_strong_2 = n_res_strong_3 = []
n_res_main_1 = n_res_main_2 = n_res_main_3 = []

N=72
K=24
L=10

#–≤—Å–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã (–∫—Ä–æ–º–µ —Ä–∞—Å—Ö–æ–¥–∞ —Ç–æ–ø–ª–∏–≤–∞) –≤ –°–ò

#–∫–æ–ª–∏—á–µ—Å—Ç–≤–æ –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å
M = 6
#–º–æ–¥—É–ª—å —É–ø—Ä—É–≥–æ—Å—Ç–∏
G = 8.3*10**10
#–∫–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ü–∏–ª–∏–Ω–¥—Ä–æ–≤
I = 6
#–¥–∏–∞–º–µ—Ç—Ä –∫–æ—Ä–µ–Ω–Ω–æ–π —à–µ–π–∫–∏
dk = 0.1
#–º–æ–º–µ–Ω—Ç –∏–Ω–µ—Ä—Ü–∏–∏ –º–∞—Ö–æ–≤–∏–∫–∞ –¥–µ–º–ø—Ñ–µ—Ä–∞
J_M = 0.01
Jp = w_s = w_01 = w_nom = 0
#—Ä–∞–¥–∏—É—Å –∫—Ä–∏–≤–æ—à–∏–ø–∞
R = 0.1
#–¥–∏–∞–º–µ—Ç—Ä —Ü–∏–ª–∏–Ω–¥—Ä–∞
D = 0.1
#–ø–∞—Ä–∞–º–µ—Ç—Ä –ø–æ–¥–æ–±–∏—è
lambd = 1.0
#–Ω–æ–º–∏–Ω–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è
n_nom = 10000
#–º–∞—Å—Å–∞ –≤–æ–∑–≤—Ä–∞—Ç–Ω–æ-–ø–æ—Å—Ç—É–ø–∞—Ç–µ–ª—å–Ω–æ –¥–≤–∏–∂—É—â–∏—Ö—Å—è —á–∞—Å—Ç–µ–π
Mvp = 0.7
#–º–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è
n_min = 6000
#–º–∞–∫—Å–∏–º–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è
n_max = 12000
#—Ç–∞–∫—Ç–Ω–æ—Å—Ç—å
tau = 4
#–æ–±—ä–µ–º —Ü–∏–ª–∏–Ω–¥—Ä–∞
V = 0.0005
#–Ω–æ–º–∏–Ω–∞–ª—å–Ω–∞—è –º–æ—â–Ω–æ—Å—Ç—å
Ne_nom = 100000.0
#–∫–æ—ç—Ñ—Ñ–∏—Ü–∏–µ–Ω—Ç –¥–µ–º–ø—Ñ–∏—Ä–æ–≤–∞–Ω–∏—è
eps = 5.0
#–Ω–æ–º–∏–Ω–∞–ª—å–Ω—ã–π —É–¥–µ–ª—å–Ω—ã–π —Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞ (–≤ –≥/(–∫–í—Ç*—á))
ge_nom=2.0

#–∫–æ–¥ –¥–ª—è –≤—ã–≤–æ–¥–∞ ascii-—Ç–∞–±–ª–∏—Ü

def params_to_path(file_path):
    # –ó–¥–µ—Å—å –∫–æ–¥ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤ –≤ –≤—ã–±—Ä–∞–Ω–Ω—ã–π —Ñ–∞–π–ª
        names_list=['–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å M',\
              '–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ü–∏–ª–∏–Ω–¥—Ä–æ–≤ I',\
              '–ú–æ–¥—É–ª—å —Å–¥–≤–∏–≥–∞ G (–ì–ü–∞)',\
              '–î–∏–∞–º–µ—Ç—Ä –∫–æ—Ä–µ–Ω–Ω–æ–π —à–µ–π–∫–∏ d (–º)',\
              '–ú–æ–º–µ–Ω—Ç –∏–Ω–µ—Ä—Ü–∏–∏ –º–∞—Ö–æ–≤–∏–∫–∞ –¥–µ–º–ø—Ñ–µ—Ä–∞ J_M (–∫–≥¬∑–º¬≤)', \
              '–†–∞–¥–∏—É—Å –∫—Ä–∏–≤–æ—à–∏–ø–∞ R (–º)',\
              '–î–∏–∞–º–µ—Ç—Ä —Ü–∏–ª–∏–Ω–¥—Ä–∞ D (–º)',\
              '–ü–∞—Ä–∞–º–µ—Ç—Ä –ø–æ–¥–æ–±–∏—è Œª',\
              '–ù–æ–º–∏–Ω–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è (–æ–±/–º)',\
              '–ú–∞—Å—Å–∞ –¥–≤–∏–∂—É—â–∏—Ö—Å—è —á–∞—Å—Ç–µ–π Mvp (–∫–≥)',\
              '–ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è (–æ–±/–º)',\
              '–ú–∞–∫—Å–∏–º–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è (–æ–±/–º)',\
              '–û–±—ä–µ–º —Ü–∏–ª–∏–Ω–¥—Ä–∞ V (–ª)',\
              '–¢–∞–∫—Ç–Ω–æ—Å—Ç—å ùúè',\
              '–ù–æ–º–∏–Ω–∞–ª—å–Ω–∞—è –º–æ—â–Ω–æ—Å—Ç—å Ne_nom (–∫–í—Ç)',\
              '–ö–æ—ç—Ñ—Ñ–∏—Ü–∏–µ–Ω—Ç –¥–µ–º–ø—Ñ–∏—Ä–æ–≤–∞–Ω–∏—è Œµ',\
              '–ù–æ–º–∏–Ω–∞–ª—å–Ω—ã–π —Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞ –≤ –≥/(–∫–í—Ç¬∑—á)',\
              '–¢–∏–ø –¥–≤–∏–≥–∞—Ç–µ–ª—è',\
              '–ö–æ–º–ø–æ–Ω–æ–≤–æ—á–Ω–∞—è —Å—Ö–µ–º–∞'  
              ]  
        entry_list=[entry_M.get(),\
	entry_I.get(),\
	entry_G.get(),\
	entry_dk.get(),\
	entry_JM.get(),\
	entry_R.get(),\
	entry_D.get(),\
	entry_lambd.get(),\
	entry_n_nom.get(),\
	entry_Mvp.get(),\
	entry_n_min.get(),\
	entry_n_max.get(),\
	entry_V.get(),\
	entry_tau.get(),\
	entry_Ne_nom.get(),\
	entry_eps.get(),\
	entry_ge_nom.get(),\
        engine_box.get(),\
        scheme_box.get()]
        
        datalist=list(zip(names_list, entry_list))

        workbook_params = Workbook()
        sheet = workbook_params.active

        # Write data from the two-dimensional list to the Excel file
        for i, row in enumerate(datalist, start=1):
            for j, value in enumerate(row, start=1):
                sheet.cell(row=i, column=j, value=value)

        # Adjust column widths based on the longest value
        for column in sheet.columns:
            max_width = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_width:
                        max_width = length
            sheet.column_dimensions[column_letter].width = max_width

        # Save the Excel file
        workbook_params.save(file_path)

def save_parameters():
    '''
    # –§—É–Ω–∫—Ü–∏—è –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤  
    from tkinter import filedialog

    # –ó–∞–ø—Ä–æ—Å–∏—Ç—å –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è –≤—ã–±—Ä–∞—Ç—å –º–µ—Å—Ç–æ –¥–ª—è —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è —Ñ–∞–π–ª–∞
    file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("–§–∞–π–ª Excel", (".xlsx"))])

    # –î–æ–±–∞–≤–∏—Ç—å –Ω—É–∂–Ω–æ–µ —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ –∫ –∏–º–µ–Ω–∏ —Ñ–∞–π–ª–∞, –≤ –∑–∞–≤–∏—Å–∏–º–æ—Å—Ç–∏ –æ—Ç –≤—ã–±–æ—Ä–∞ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—è
    if file_path:
        if file_path.endswith(".xls"):
            # –ï—Å–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –≤—ã–±—Ä–∞–ª —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ xls
            file_path += ".xls"
        else:
            # –ï—Å–ª–∏ –ø–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –≤—ã–±—Ä–∞–ª —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ xlsx –∏–ª–∏ –Ω–µ —É–∫–∞–∑–∞–ª —Ä–∞—Å—à–∏—Ä–µ–Ω–∏–µ
            file_path += ".xlsx"

    # –ü—Ä–æ–≤–µ—Ä–∏—Ç—å, —á—Ç–æ file_path –æ–∫–∞–Ω—á–∏–≤–∞–µ—Ç—Å—è –Ω–∞ .xls –∏–ª–∏ .xlsx
    if file_path and (file_path.endswith(".xls") or file_path.endswith(".xlsx")):
        params_to_path(file_path)
        pass
    '''
    initial_filename = "new_params.xlsx"
        
    file_path = filedialog.asksaveasfilename(
        initialfile=initial_filename,    
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx"), ("Excel 97-2003 files", "*.xls")],
        title="–°–æ—Ö—Ä–∞–Ω–∏—Ç—å —Ñ–∞–π–ª –∫–∞–∫"
    )
    if file_path and (file_path.endswith(".xls") or file_path.endswith(".xlsx")):
        params_to_path(file_path)
        pass
    
    '''
    filetypes = (
        ('Excel files', '*xlsx'),
        ('Excel files', '*xls')
    )
    file_path = asksaveasfilename(defaultextension = ".xls", filetypes=filetypes)
    if file_path and (file_path.endswith(".xls") or file_path.endswith(".xlsx")):
        params_to_path(file_path)
        pass
    else:
        tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ù–µ–≤–µ—Ä–Ω—ã–π —Ñ–æ—Ä–º–∞—Ç —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è")
    '''
    
  

   
def load_parameters():
    # –§—É–Ω–∫—Ü–∏—è –¥–ª—è –∑–∞–≥—Ä—É–∑–∫–∏ –ø–∞—Ä–∞–º–µ—Ç—Ä–æ–≤
    filetypes = (
        ('Excel files', '*xlsx'),
        ('Excel files', '*xls')
    )
    
    file_path = askopenfilename(filetypes=filetypes)
    if file_path:
        load_params(file_path)
        pass

def load_params(file_path):
    df = read_excel(file_path, header = None)

    # –ü–æ–ª—É—á–µ–Ω–∏–µ –≤—Ç–æ—Ä–æ–π –∫–æ–ª–æ–Ω–∫–∏ –ø–æ –∏–Ω–¥–µ–∫—Å—É
    second_column = df.iloc[:, 1].astype(str).tolist()
    #print(second_column)
    if params_checked(second_column):
        load_entries(second_column)
    else:
        tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ü–∞—Ä–∞–º–µ—Ç—Ä—ã –Ω–µ –ø–æ–ª–Ω—ã –∏–ª–∏ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã")

def params_checked(params):
    try: 
        q=np.zeros(19)
        q[0]=is_number(params[0])
        q[1]=is_number(params[1])
        q[2]=is_float(params[2]) or is_float(params[2].replace(',','.'))
        q[3]=is_float(params[3]) or is_float(params[3].replace(',','.'))
        q[4]=is_float(params[4]) or is_float(params[4].replace(',','.'))
        q[5]=is_float(params[5]) or is_float(params[5].replace(',','.'))
        q[6]=is_float(params[6]) or is_float(params[6].replace(',','.'))
        q[7]=is_float(params[7]) or is_float(params[7].replace(',','.'))
        q[8]=is_number(params[8])
        q[9]=is_float(params[9]) or is_float(params[9].replace(',','.'))
        q[10]=is_number(params[10])
        q[11]=is_number(params[11])
        q[12]=is_float(params[12]) or is_float(params[12].replace(',','.'))
        q[13]=is_number(params[13])
        q[14]=is_float(params[14]) or is_float(params[14].replace(',','.'))
        q[15]=is_float(params[15]) or is_float(params[15].replace(',','.'))
        q[16]=is_float(params[16]) or is_float(params[16].replace(',','.'))
        q[17]=params[17] in ['–ò—Å–∫—Ä–æ–≤–æ–π', '–î–∏–∑–µ–ª—å–Ω—ã–π']
        q[18]=params[18] in ['R', 'V', 'O']
    except:
        return 0
    
    #print(q.prod())
    return q.prod()

def load_entries(params):
    entry_list=[entry_M,\
	entry_I,\
	entry_G,\
	entry_dk,\
	entry_JM,\
	entry_R,\
	entry_D,\
	entry_lambd,\
	entry_n_nom,\
	entry_Mvp,\
	entry_n_min,\
	entry_n_max,\
	entry_V,\
	entry_tau,\
	entry_Ne_nom,\
	entry_eps,\
	entry_ge_nom]
    for i in range(17):
        entry_list[i].delete(0, tk.END)
        #print(params[i].replace('.',','))
        entry_list[i].insert(0, params[i].replace('.',','))
        
    engine_box.set(params[17])
    scheme_box.set(params[18])
    

def print_pretty_table(data, cell_sep=' | ', header_separator=True):
    rows = len(data)
    cols = len(data[0])

    col_width = []
    for col in range(cols):
        columns = [data[row][col] for row in range(rows)]
        col_width.append(len(max(columns, key=len)))

    separator = "-+-".join('-' * n for n in col_width)

    for i, row in enumerate(range(rows)):
        if i == 1 and header_separator:
            print(separator)

        result = []
        for col in range(cols):
            item = data[row][col].rjust(col_width[col])
            result.append(item)

        print(cell_sep.join(result))
    print('\n')

def write_pretty_table_to_txt(data, filename, cell_sep=' | ', header_separator=True):
    rows = len(data)
    cols = len(data[0])

    col_width = []
    for col in range(cols):
        columns = [data[row][col] for row in range(rows)]
        col_width.append(len(max(columns, key=len)))

    separator = "-+-".join('-' * n for n in col_width)

    with open(filename, 'a') as file:
        for i, row in enumerate(range(rows)):
            if i == 1 and header_separator:
                file.write(separator + '\n')

            result = []
            for col in range(cols):
                item = data[row][col].rjust(col_width[col])
                result.append(item)

            file.write(cell_sep.join(result) + '\n')
        file.write('\n')

def write_to_file(data, filename):
    with open(filename, 'a') as file:
        file.write(data)

def is_number(val):
    try:
        int(val)
        f = True
    except:
        f = False
    return f

def is_float(val):
    try:
        float(val)
        f = True
    except:
        f = False
    return f 

#–ø—Ä–æ–≤–µ—Ä–∫–∞ –≤–≤–æ–¥–∏–º–æ–≥–æ –∑–Ω–∞—á–µ–Ω–∏—è –Ω–∞ —Ü–µ–ª–æ–µ —á–∏—Å–ª–æ
def is_valid_number(newval):
    return newval == '' or newval.isnumeric()

#–ø—Ä–æ–≤–µ—Ä–∫–∞ –≤–≤–æ–¥–∏–º–æ–≥–æ –∑–Ω–∞—á–µ–Ω–∏—è –Ω–∞ –≤–µ—â–µ—Å—Ç–≤–µ–Ω–Ω–æ–µ —á–∏—Å–ª–æ
def is_valid_float(newval):
    try:
        float(newval.replace('.','a').replace(',','.'))
        f = True
    except:
        f = False
    return newval == '' or f

#–ø—Ä–æ–≤–µ—Ä–∫–∞ –Ω–∞ –∑–∞–ø–æ–ª–Ω–µ–Ω–∏–µ –≤—Å–µ—Ö —è—á–µ–µ–∫
def is_fulled():
    entry_list=[entry_M,\
    entry_I,\
    entry_G,\
    entry_dk,\
    entry_JM,\
    entry_R,\
    entry_D,\
    entry_lambd,\
    entry_n_nom,\
    entry_Mvp,\
    entry_n_min,\
    entry_n_max,\
    entry_V,\
    entry_tau,\
    entry_Ne_nom,\
    entry_eps,\
    entry_ge_nom]
    for entry in entry_list:
        if entry.get()=='':
            raise ValueError("–ó–∞–ø–æ–ª–Ω–µ–Ω—ã –Ω–µ –≤—Å–µ —è—á–µ–π–∫–∏!")

#—Ñ—É–Ω–∫—Ü–∏—è —Ñ–æ—Ä–º–∏—Ä—É–µ—Ç –º–∞—Å—Å–∏–≤ a[i]
def find_a(w):
    global a
    a=np.ones(M)
    for i in range(1, M):
      aJ=a*J
      a[i]=a[i-1]-w**2*np.sum(aJ[0:i])/C[i-1]
    return a

#–º–∏–Ω–∏–º–∏–∑–∏—Ä—É–µ–º–∞—è –æ—à–∏–±–∫–∞
def min_to_find(w):
    Ja=J*find_a(w)
    return abs(np.sum(Ja)/(np.max(abs(Ja))))

#–ø–æ–∏—Å–∫ —Å–æ–±—Å—Ç–≤–µ–Ω–Ω–æ–π —á–∞—Å—Ç–æ—Ç—ã
def find_w_s():
    global w_s
    w_s = w_01
    while (w_s>0 and min_to_find(w_s)>0.05):
        w_s = w_s-w_01*0.01

    if min_to_find(w_s)>0.05:
          while (w_s<10*w_01 and min_to_find(w_s)>0.05):
              w_s = w_s+w_01*0.01
    if min_to_find(w_s)>0.05:
        tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ù–µ –Ω–∞–π–¥–µ–Ω–∞ —Å–æ–±—Å—Ç–≤–µ–Ω–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞")
        raise ValueError()
            

def print_Jp():
    #Jp_print=Jp*10**8
    text='–ü–æ–ª—è—Ä–Ω—ã–π –º–æ–º–µ–Ω—Ç –∏–Ω–µ—Ä—Ü–∏–∏ Jp = ' + ("%.8f" %Jp)+' –º\u2074'
    text=text.replace('.',',')
    Jp_label.config(text=text)
    Jp_label.grid(column=4, row=0, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('–ü–æ–ª—è—Ä–Ω—ã–π –º–æ–º–µ–Ω—Ç –∏–Ω–µ—Ä—Ü–∏–∏ Jp = ' + ("%.8f" %Jp)+' –º^4\n', f_out)
        

def print_Ji():
    global wb
    str1=['i']+list(map(str, range(1,M+1)))
    str2=['J_i,–∫–≥*–º^2']+list(map(str, J.tolist()))
    data=[str1,str2]

    print('–ú–æ–º–µ–Ω—Ç—ã –∏–Ω–µ—Ä—Ü–∏–∏ –º–∞—Å—Å –≤ –∫–≥*–º^2\n\n')
    print_pretty_table(data)
    
    if textfile_on.get():
        
        write_to_file('–ú–æ–º–µ–Ω—Ç—ã –∏–Ω–µ—Ä—Ü–∏–∏ –º–∞—Å—Å –≤ –∫–≥*–º^2\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)
        
    if excelfile_on.get():
        str1=['i']+list(map(str, range(1,M+1)))
        str2=['J_i, –∫–≥\u00B7–º\u00B2']+list(map(str, J.tolist()))
        data=[str1,str2]
        
        df = DataFrame(data[1:], columns=data[0])
        wb = Workbook()
        ws = wb.active
        ws.title = '–ú–æ–º–µ–Ω—Ç—ã –∏–Ω–µ—Ä—Ü–∏–∏ Ji'

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')

def print_JlC():
    global wb
    
    str1=['i,i+1']
    for i in range(1,M):
        nums=str(i)+','+str(i+1)
        str1.append(nums)
    C_print=C.tolist()
    l_print=l.tolist()
    C_print=[str(round(num)) for num in C]
    l_print=["{:.3f}".format(num) for num in l]
  
    str2=['C_i,i+1, –ù*–º/—Ä–∞–¥']+C_print
    str3=['l_i,i+1, –º']+l_print
    data=[str1,str2,str3]

    print('–ñ–µ—Å—Ç–∫–æ—Å—Ç–∏ –∏ –¥–ª–∏–Ω—ã –ø—Ä–æ–º–µ–∂—É—Ç–∫–æ–≤\n\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–ñ–µ—Å—Ç–∫–æ—Å—Ç–∏ –∏ –¥–ª–∏–Ω—ã –ø—Ä–æ–º–µ–∂—É—Ç–∫–æ–≤\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=['i,i+1']
        for i in range(1,M):
            nums=str(i)+','+str(i+1)
            str1.append(nums)
        C_print=C.tolist()
        l_print=l.tolist()
        C_print=[str(round(num)) for num in C]
        l_print=["{:.3f}".format(num) for num in l]
      
        str2=['C_i,i+1, –ù\u00B7–º/—Ä–∞–¥']+C_print
        str3=['l_i,i+1, –º']+l_print
        data=[str1,str2,str3]
        
        ws = wb.create_sheet(title='–î–ª–∏–Ω—ã –∏ –∂–µ—Å—Ç–∫–æ—Å—Ç–∏')
        data=[str1,str2,str3]
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')

def print_w_s():
    w_s_label.config(text='–°–æ–±—Å—Ç–≤–µ–Ω–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ w_s = ' + ("%.0f" %w_s)+' —Ä–∞–¥/—Å')
    w_s_label.grid(column=4, row=1, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('–°–æ–±—Å—Ç–≤–µ–Ω–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ w_s = ' + ("%.8f" %w_s)+' —Ä–∞–¥/—Å\n', f_out)

def print_a():
    global wb
    str1=['i']+list(map(str, range(1,M+1)))
    str2=['a_i']+list(map(str, (np.round(a,3)).tolist()))
    data=[str1, str2]
    
    print('–ö–æ—ç—Ñ—Ñ–∏—Ü–∏–µ–Ω—Ç—ã a_i\n\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–ö–æ—ç—Ñ—Ñ–∏—Ü–∏–µ–Ω—Ç—ã a_i\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():     
        ws = wb.create_sheet(title='–ö–æ—ç—Ñ—Ñ–∏—Ü–∏–µ–Ω—Ç—ã a_i')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')

def print_n_res():
    global wb
    data=[['K_m', 'n_res, –æ–±/–º']]
    
    data.append(['K_m_strong_1']+[str(round(n_res_strong_1))])
    data.append(['K_m_main_1']+[str(round(n_res_main_1))])

    data.append(['K_m_strong_2']+[str(round(n_res_strong_2))])
    data.append(['K_m_main_2']+[str(round(n_res_main_2))])

    data.append(['K_m_strong_3']+[str(round(n_res_strong_3))])
    data.append(['K_m_main_3']+[str(round(n_res_main_3))])

    print('–°–∏–ª—å–Ω—ã–µ –∏ –≥–ª–∞–≤–Ω—ã–µ –≥–∞—Ä–º–æ–Ω–∏–∫–∏\n\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–°–∏–ª—å–Ω—ã–µ –∏ –≥–ª–∞–≤–Ω—ã–µ –≥–∞—Ä–º–æ–Ω–∏–∫–∏\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)
    if excelfile_on.get():     
        ws = wb.create_sheet(title='–°–∏–ª—å–Ω—ã–µ –∏ –≥–ª–∞–≤–Ω—ã–µ –≥–∞—Ä–º–æ–Ω–∏–∫–∏')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_M_rot_g_mean():
    text='–°—Ä–µ–¥–Ω–∏–π –∫—Ä—É—Ç—è—â–∏–π –≥–∞–∑–æ–≤—ã–π –º–æ–º–µ–Ω—Ç = ' + ("%.2f" %M_rot_g_mean)+' –ù\u00B7–º'
    text=text.replace('.',',')
    M_rot_g_mean_label.config(text=text)
    M_rot_g_mean_label.grid(column=4, row=2, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('–°—Ä–µ–¥–Ω–∏–π –∫—Ä—É—Ç—è—â–∏–π –≥–∞–∑–æ–≤—ã–π –º–æ–º–µ–Ω—Ç = ' + ("%.2f" %M_rot_g_mean)+' –ù*–º\n', f_out)

def print_M_rot_j_mean():
    text='–°—Ä–µ–¥–Ω–∏–π –∫—Ä—É—Ç—è—â–∏–π –≥–∞–∑–æ–≤—ã–π –º–æ–º–µ–Ω—Ç = ' + ("%.2f" %M_rot_g_mean)+' –ù\u00B7–º'
    text=text.replace('.',',')
    M_rot_j_mean_label.config(text=text)
    M_rot_j_mean_label.grid(column=4, row=3, sticky=tk.W, padx=10)
    #print(textfile_on.get())
    if textfile_on.get():
        write_to_file('–°—Ä–µ–¥–Ω–∏–π –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã–π –≥–∞–∑–æ–≤—ã–π –º–æ–º–µ–Ω—Ç = ' + ("%.2f" %M_rot_j_mean)+' –ù*–º\n', f_out)

def print_Fourier_table():
    global wb
    data=[['k', '–ú_–∫—Ä_–≥–∞–∑_k, –ù*–º', 'phi_–≥–∞–∑_k', '–ú_–∫—Ä_–∏–Ω–µ—Ä—Ü_k, –ù*–º', 'phi_–∏–Ω–µ—Ä—Ü_k']]
    for k in range(K):
        data.append([str(k+1), str(round(M_rot_g_sqr[k],2)), str(round(phi_g[k],2)), str(round(M_rot_j_sqr[k],2)), str(round(phi_j[k],2))])
    print('–ú–æ–º–µ–Ω—Ç—ã –≥–∞—Ä–º–æ–Ω–∏–∫\n\n')
    print_pretty_table(data)
    
    if textfile_on.get():
        write_to_file('–ú–æ–º–µ–Ω—Ç—ã –≥–∞—Ä–º–æ–Ω–∏–∫\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        ws = wb.create_sheet(title='–ú–æ–º–µ–Ω—Ç—ã –≥–∞—Ä–º–æ–Ω–∏–∫')
        
        data[0]=['k', '–ú_–∫—Ä_–≥–∞–∑_k, –ù\u00B7–º', 'phi_–≥–∞–∑_k', '–ú_–∫—Ä_–∏–Ω–µ—Ä—Ü_k, –ù\u00B7–º', 'phi_–∏–Ω–µ—Ä—Ü_k']
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        #wb.save('output.xlsx')
    

def print_M_rot_n_g():
    global wb
    str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(M_rot_n_g_strong_1).astype(int).astype(str).tolist()]
    str3=[['K_main_1']+np.round(M_rot_n_g_main_1).astype(int).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(M_rot_n_g_strong_2).astype(int).astype(str).tolist()]
    str5=[['K_main_2']+np.round(M_rot_n_g_main_2).astype(int).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(M_rot_n_g_strong_3).astype(int).astype(str).tolist()]
    str7=[['K_main_3']+np.round(M_rot_n_g_main_3).astype(int).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('–ú–æ–º–µ–Ω—Ç—ã –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫ –≤ –ù*–º\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–ú–æ–º–µ–Ω—Ç—ã –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫ –≤ –ù*–º\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_g_strong_1).astype(int).astype(str).tolist()]
        str3=[['–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_g_main_1).astype(int).astype(str).tolist()]
        str4=[['–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_g_strong_2).astype(int).astype(str).tolist()]
        str5=[['–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_g_main_2).astype(int).astype(str).tolist()]
        str6=[['–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_g_strong_3).astype(int).astype(str).tolist()]
        str7=[['–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_g_main_3).astype(int).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7
        ws = wb.create_sheet(title='–ú–æ–º–µ–Ω—Ç—ã –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_M_rot_n_j():
    global wb
    str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(M_rot_n_j_strong_1).astype(int).astype(str).tolist()]
    str3=[['K_main_1']+np.round(M_rot_n_j_main_1).astype(int).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(M_rot_n_j_strong_2).astype(int).astype(str).tolist()]
    str5=[['K_main_2']+np.round(M_rot_n_j_main_2).astype(int).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(M_rot_n_j_strong_3).astype(int).astype(str).tolist()]
    str7=[['K_main_3']+np.round(M_rot_n_j_main_3).astype(int).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('–ú–æ–º–µ–Ω—Ç—ã –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫ –≤ –ù*–º\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–ú–æ–º–µ–Ω—Ç—ã –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫ –≤ –ù*–º\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_j_strong_1).astype(int).astype(str).tolist()]
        str3=[['–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_j_main_1).astype(int).astype(str).tolist()]
        str4=[['–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_j_strong_2).astype(int).astype(str).tolist()]
        str5=[['–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_j_main_2).astype(int).astype(str).tolist()]
        str6=[['–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_j_strong_3).astype(int).astype(str).tolist()]
        str7=[['–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è, –ù\u00B7–º']+np.round(M_rot_n_j_main_3).astype(int).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7
        ws = wb.create_sheet(title='–ú–æ–º–µ–Ω—Ç—ã –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_Ai(i):
    str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(A_strong_1[i],5).astype(float).astype(str).tolist()]
    str3=[['K_main_1']+np.round(A_main_1[i],5).astype(float).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(A_strong_2[i],5).astype(float).astype(str).tolist()]
    str5=[['K_main_2']+np.round(A_main_2[i],5).astype(float).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(A_strong_3[i],5).astype(float).astype(str).tolist()]
    str7=[['K_main_3']+np.round(A_main_3[i],5).astype(float).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('–ê–º–ø–ª–∏—Ç—É–¥–∞ –∫–æ–ª–µ–±–∞–Ω–∏–π –¥–ª—è i =',str(i+1),'–≤ –º\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–ê–º–ø–ª–∏—Ç—É–¥–∞ –∫–æ–ª–µ–±–∞–Ω–∏–π –¥–ª—è i = '+str(i+1)+' –≤ –º\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

def print_excel_Ai():
    data=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    
    for i in range(M):
        str1=[['i = '+str(i+1)]+['']*(L)]
        str2=[['–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è, –º']+np.round(A_strong_1[i],5).astype(float).astype(str).tolist()]
        str3=[['–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è –º']+np.round(A_main_1[i],5).astype(float).astype(str).tolist()]
        str4=[['–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è –º']+np.round(A_strong_2[i],5).astype(float).astype(str).tolist()]
        str5=[['–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è –º']+np.round(A_main_2[i],5).astype(float).astype(str).tolist()]
        str6=[['–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è –º']+np.round(A_strong_3[i],5).astype(float).astype(str).tolist()]
        str7=[['–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è –º']+np.round(A_main_3[i],5).astype(float).astype(str).tolist()]
        str8=[['']*(L+1)]
        data+=str1+str2+str3+str4+str5+str6+str7+str8

    ws = wb.create_sheet(title='–ê–º–ø–ª–∏—Ç—É–¥—ã –∫–æ–ª–µ–±–∞–Ω–∏–π')
        
    df = DataFrame(data[1:], columns=data[0])

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
    #wb.save('output.xlsx')
    
def print_M_upr(i):
    str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(M_upr_strong_1[i],1).astype(float).astype(str).tolist()]
    str3=[['K_main_1']+np.round(M_upr_main_1[i],1).astype(float).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(M_upr_strong_2[i],1).astype(float).astype(str).tolist()]
    str5=[['K_main_2']+np.round(M_upr_main_2[i],1).astype(float).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(M_upr_strong_3[i],1).astype(float).astype(str).tolist()]
    str7=[['K_main_3']+np.round(M_upr_main_3[i],1).astype(float).astype(str).tolist()]
    data=str1+str2+str3+str4+str5+str6+str7

    print('–£–ø—Ä—É–≥–∏–µ –º–æ–º–µ–Ω—Ç—ã –¥–ª—è i =',str(i+1),'–≤ –ù*–º\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–£–ø—Ä—É–≥–∏–µ –º–æ–º–µ–Ω—Ç—ã –¥–ª—è i = '+str(i+1)+' –≤ –ù*–º\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

def print_excel_M_upr():
    data=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    
    for i in range(M-1):
        str1=[['i, i+1 = '+str(i+1)+', '+str(i+2)]+['']*(L)]
        str2=[['–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è, –º']+np.round(M_upr_strong_1[i],5).astype(float).astype(str).tolist()]
        str3=[['–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è –º']+np.round(M_upr_main_1[i],5).astype(float).astype(str).tolist()]
        str4=[['–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è –º']+np.round(M_upr_strong_2[i],5).astype(float).astype(str).tolist()]
        str5=[['–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è –º']+np.round(M_upr_main_2[i],5).astype(float).astype(str).tolist()]
        str6=[['–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è –º']+np.round(M_upr_strong_3[i],5).astype(float).astype(str).tolist()]
        str7=[['–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è –º']+np.round(M_upr_main_3[i],5).astype(float).astype(str).tolist()]
        str8=[['']*(L+1)]
        data+=str1+str2+str3+str4+str5+str6+str7+str8

    ws = wb.create_sheet(title='–£–ø—Ä—É–≥–∏–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è')
        
    df = DataFrame(data[1:], columns=data[0])

    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font

    for column_cells in ws.columns:
        max_length = 0
        for cell in column_cells:
            if len(str(cell.value)) > max_length:
                max_length = len(cell.value)
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
    #wb.save('output.xlsx')

def print_T():
    T_label.config(text='–ú–∞–∫—Å–∏–º–∞–ª—å–Ω–æ–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ T = ' + ("%.0f" %T)+' –ü–∞')
    T_label.grid(column=4, row=4, sticky=tk.W, padx=10)
    if textfile_on.get():
        write_to_file('–ú–∞–∫—Å–∏–º–∞–ª—å–Ω–æ–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ T = ' + ("%.0f" %T)+' –ü–∞\n', f_out)

def print_N():
    str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(N_strong_1).astype(int).astype(str).tolist()]
    str3=[['K_main_1']+np.round(N_main_1).astype(int).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(N_strong_2).astype(int).astype(str).tolist()]
    str5=[['K_main_2']+np.round(N_main_2).astype(int).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(N_strong_3).astype(int).astype(str).tolist()]
    str7=[['K_main_3']+np.round(N_main_3).astype(int).astype(str).tolist()]
    str8=[['–°—É–º–º–∞']+np.round(N_sum).astype(int).astype(str).tolist()]
    
    data=str1+str2+str3+str4+str5+str6+str7+str8

    print('–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏ –≤ –í—Ç\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏ –≤ –í—Ç\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)
    if excelfile_on.get():
        str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è, –í—Ç']+np.round(N_strong_1).astype(int).astype(str).tolist()]
        str3=[['–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è, –í—Ç']+np.round(N_main_1).astype(int).astype(str).tolist()]
        str4=[['–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è, –í—Ç']+np.round(N_strong_2).astype(int).astype(str).tolist()]
        str5=[['–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è, –í—Ç']+np.round(N_main_2).astype(int).astype(str).tolist()]
        str6=[['–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è, –í—Ç']+np.round(N_strong_3).astype(int).astype(str).tolist()]
        str7=[['–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è, –í—Ç']+np.round(N_main_3).astype(int).astype(str).tolist()]
        str8=[['–°—É–º–º–∞—Ä–Ω–∞—è, –í—Ç']+np.round(N_sum).astype(int).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7+str8
        ws = wb.create_sheet(title='–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width
            
        #wb.save('output.xlsx')

def print_dg():
    str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
    str2=[['K_strong_1']+np.round(dg_strong_1,3).astype(float).astype(str).tolist()]
    str3=[['K_main_1']+np.round(dg_main_1,3).astype(float).astype(str).tolist()]
    str4=[['K_strong_2']+np.round(dg_strong_2,3).astype(float).astype(str).tolist()]
    str5=[['K_main_2']+np.round(dg_main_2,3).astype(float).astype(str).tolist()]
    str6=[['K_strong_3']+np.round(dg_strong_3,3).astype(float).astype(str).tolist()]
    str7=[['K_main_3']+np.round(dg_main_3,3).astype(float).astype(str).tolist()]
    str8=[['–°—É–º–º–∞']+np.round(dg_sum,3).astype(float).astype(str).tolist()]
    
    data=str1+str2+str3+str4+str5+str6+str7+str8

    print('–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞ –≤ –≥/(–∫–í—Ç*—á)\n')
    print_pretty_table(data)
    if textfile_on.get():
        write_to_file('–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞ –≤ –≥/(–∫–í—Ç*—á)\n\n', f_tables)
        write_pretty_table_to_txt(data, f_tables)

    if excelfile_on.get():
        str1=[['n, –æ–±/–º']+np.round(n_diap).astype(int).astype(str).tolist()]
        str2=[['–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è, –≥/(–∫–í—Ç\u00B7—á)']+np.round(dg_strong_1, 3).astype(float).astype(str).tolist()]
        str3=[['–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è, –≥/(–∫–í—Ç\u00B7—á)']+np.round(dg_main_1, 3).astype(float).astype(str).tolist()]
        str4=[['–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è, –≥/(–∫–í—Ç\u00B7—á)']+np.round(dg_strong_2, 3).astype(float).astype(str).tolist()]
        str5=[['–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è, –≥/(–∫–í—Ç\u00B7—á)']+np.round(dg_main_2, 3).astype(float).astype(str).tolist()]
        str6=[['–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è, –≥/(–∫–í—Ç\u00B7—á)']+np.round(dg_strong_3, 3).astype(float).astype(str).tolist()]
        str7=[['–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è, –≥/(–∫–í—Ç\u00B7—á)']+np.round(dg_main_3, 3).astype(float).astype(str).tolist()]
        str8=[['–°—É–º–º–∞—Ä–Ω—ã–π, –≥/(–∫–í—Ç\u00B7—á)']+np.round(dg_sum, 3).astype(float).astype(str).tolist()]
        data=str1+str2+str3+str4+str5+str6+str7+str8
        ws = wb.create_sheet(title='–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞')
        
        df = DataFrame(data[1:], columns=data[0])

        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)

        header_font = Font(bold=True)
        for cell in ws[1]:
            cell.font = header_font

        for column_cells in ws.columns:
            max_length = 0
            for cell in column_cells:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_cells[0].column_letter].width = adjusted_width


            
        wb.save('output.xlsx')

#–ü–æ–ø—ã—Ç–∫–∞ –∑–∞–ø—É—Å—Ç–∏—Ç—å –∏ –∫–æ—Ä—Ä–µ–∫—Ç–Ω–æ –ø—Ä–æ–≤–µ—Å—Ç–∏ —Ä–∞—Å—Å—á–µ—Ç
def try_calculate():
    try:
        is_fulled()
        calculate2()
    except ValueError as error:
        tk.messagebox.showerror("–û—à–∏–±–∫–∞", str(error))
def calculate2():
    try:
        calculate()
    except ValueError as error:
        tk.messagebox.showerror('–û—à–∏–±–∫–∞ –ø—Ä–∏ –≤—ã—á–∏—Å–ª–µ–Ω–∏—è—Ö', '–ü—Ä–æ–≤–µ—Ä—å—Ç–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã')
        
        
#–ó–∞–ø—É—Å–∫ —Ä–∞—Å—á–µ—Ç–∞
def calculate():
    global N, K, L
    global M, G, I, dk, Jp, J_M, R, D, n_nom, Mvp, n_min, n_max, V, tau, Ne_nom, eps, g_e_nom
    global J, C, l, n_diap
    global w_01, w_nom
    global j, beta, Pj, Pg, M_rot_g, M_rot_j, Ag, Bg, M_rot_g_sqr, phi_g, M_rot_g_mean
    global M_rot_n_g_strong_1, M_rot_n_g_strong_2, M_rot_n_g_strong_3
    global M_rot_n_g_main_1, M_rot_n_g_main_2, M_rot_n_g_main_3
    global M_rot_n_j_strong_1, M_rot_n_j_strong_2, M_rot_n_j_strong_3
    global M_rot_n_j_main_1, M_rot_n_j_main_2, M_rot_n_j_main_3
    global Aj, Bj, M_rot_j_sqr, phi_j, M_rot_j_mean

    global A1_strong_1, A1_strong_2, A1_strong_3
    global A1_main_1, A1_main_2, A1_main_3
    global A_strong_1, A_strong_2, A_strong_3
    global A_main_1, A_main_2, A_main_3
    
    #–≥–∞—Ä–º–æ–Ω–∏–∫–∏
    global K_m_main_1, K_m_main_2, K_m_main_3
    global K_m_strong_1, K_m_strong_2, K_m_strong_3
    #—Ä–µ–∑–æ–Ω–∞–Ω—Å–Ω—ã–µ –æ–±–æ—Ä–æ—Ç—ã
    global n_res_strong_1, n_res_strong_2, n_res_strong_3
    global n_res_main_1, n_res_main_2, n_res_main_3
    
    global phi, pg
    global a_coef, b_coef, c_coef
    global M_upr_strong_1, M_upr_strong_2, M_upr_strong_3
    global M_upr_main_1, M_upr_main_2, M_upr_main_3
    global T
        
    global N_strong_1, N_strong_2, N_strong_3
    global N_main_1, N_main_2, N_main_3
    global N_sum

    global dg_strong_1, dg_strong_2, dg_strong_3
    global dg_main_1, dg_main_2, dg_main_3
    global dg_sum
    
    #–∑–∞–¥–∞–≤–∞–µ–º—ã–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã –≤ –°–ò
    M = int(entry_M.get())
    I = int(entry_I.get())
    G = 10**9*float(entry_G.get().replace(',','.'))
    dk = float(entry_dk.get().replace(',','.'))
    J_M = float(entry_JM.get().replace(',','.'))
    R = float(entry_R.get().replace(',','.'))
    D = float(entry_D.get().replace(',','.'))
    lambd = float(entry_lambd.get().replace(',','.'))
    n_nom = int(entry_n_nom.get())
    Mvp = float(entry_Mvp.get().replace(',','.'))
    n_min = int(entry_n_min.get())
    n_max = int(entry_n_max.get())
    V = float(entry_V.get().replace(',','.'))/1000
    tau = int(entry_tau.get())
    Ne_nom = float(entry_Ne_nom.get().replace(',','.'))*1000
    eps = float(entry_eps.get().replace(',','.'))
    ge_nom = float(entry_ge_nom.get().replace(',','.'))
    
    input_file = filedialog.askopenfilename()
    if input_file=='':
        return 0
    open(input_file, "r")

    #open(f_out, "w")
    #open(f_tables, "w")

    

    a_coef=1
    b_coef=1
    c_coef=-1
    if engine_box.get()=='–î–∏–∑–µ–ª—å–Ω—ã–π':
        a_coef=0.87
        b_coef=1.13
        c_coef=-1
    
    J = np.array(read_excel(input_file, sheet_name=0, header=None).iloc[1:, 1])
    C = np.array(read_excel(input_file, sheet_name=1, header=None).iloc[1:, 1])
    if (len(J)!=M or len(C)!=(M-1)):
        #tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ù–µ—Å–æ–æ—Ç–≤–µ—Ç—Å—Ç–≤–∏–µ –∫–æ–ª–∏—á–µ—Å—Ç–≤—É –º–∞—Å—Å –≤ –æ—Ç–∫—Ä—ã—Ç–æ–º —Ñ–∞–π–ª–µ")
        raise ValueError()
        
    print('\n')
    #print(J)
    #print(C)
    
    #1. –≠–∫–≤–∏–≤–∞–ª–µ–Ω—Ç–Ω–∞—è —Ä–∞—Å—á–µ—Ç–Ω–∞—è —Å—Ö–µ–º–∞
    
    Jp = np.pi*dk**4/32
    #–∑–¥–µ—Å—å –±—É–¥—É—Ç –¥–ª–∏–Ω—ã l_i,i+1 –º–µ–∂–¥—É –º–∞—Å—Å–∞–º–∏
    l = np.zeros(M-1); 
    #–≤—ã—á–∏—Å–ª—è–µ–º –¥–ª–∏–Ω—ã
    l = G*Jp/C

    #3. –°–æ–±—Å—Ç–≤–µ–Ω–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ —ç–∫–≤–∏–≤–∞–ª–µ–Ω—Ç–Ω–æ–π —Å—Ö–µ–º—ã
    
    try:
       J_sum = np.sum(J[0:M-1])
    except:
       #tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ü–∞—Ä–∞–º–µ—Ç—Ä—ã –Ω–µ –ø–æ–ª–Ω—ã –∏–ª–∏ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã")
       raise ValueError()
    try:    
       l_sum = np.sum(l[0:M-2])+l[M-2]
    except:
       #tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ü–∞—Ä–∞–º–µ—Ç—Ä—ã –Ω–µ –ø–æ–ª–Ω—ã –∏–ª–∏ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã")
       raise ValueError()   
    C_sum = G*Jp/l_sum

    #–ø—Ä–∏–±–ª–∏–∂–µ–Ω–Ω–æ–µ –∑–Ω–∞—á–µ–Ω–∏–µ —á–∞—Å—Ç–æ—Ç—ã
    w_01 = np.sqrt(C_sum*(J_sum+J_M)/(J_sum*J_M))
    find_w_s()

    #4. –ì–ª–∞–≤–Ω—ã–µ –∏ —Å–∏–ª—å–Ω—ã–µ –≥–∞—Ä–º–æ–Ω–∏–∫–∏
    
    K_m_main_1 = I*1; K_m_main_2 = I*2; K_m_main_3 = I*3; 
    K_m_strong_1 = int(0.5*I*1); K_m_strong_2 = int(0.5*I*2); K_m_strong_3 = int(0.5*I*3);

    #5. –†–µ–∑–æ–Ω–∞–Ω—Å–Ω—ã–µ –æ–±–æ—Ä–æ—Ç—ã

    n_res_strong_1 = 30*w_s/(np.pi*K_m_strong_1)
    n_res_main_1 = 30*w_s/(np.pi*K_m_main_1)

    n_res_strong_2 = 30*w_s/(np.pi*K_m_strong_2)
    n_res_main_2 = 30*w_s/(np.pi*K_m_main_2)

    n_res_strong_3 = 30*w_s/(np.pi*K_m_strong_3)
    n_res_main_3 = 30*w_s/(np.pi*K_m_main_3)

    #6. –ì–∞—Ä–º–æ–Ω–∏—á–µ—Å–∫–∏–π –∞–Ω–∞–ª–∏–∑

    #–ø–µ—Ä–µ–≤–æ–¥–∏–º –≤ —Ä–∞–¥–∏–∞–Ω—ã!
    phi = (np.pi/180)*(np.array(read_excel(input_file, sheet_name=2, header=None).iloc[1:, 0]))
    
    pg = np.array(read_excel(input_file, sheet_name=2, header=None).iloc[1:, 1])
    
    w_nom = 2*np.pi*n_nom/60
    phi=np.array(phi.tolist())
    
    j = R*w_nom**2*(np.cos(phi)+lambd*np.cos(2*phi))
    
    alph=lambd*np.sin(phi)
    for i in range(len(alph)):
        if abs(alph[i])>1:
            #tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ü–∞—Ä–∞–º–µ—Ç—Ä—ã –Ω–µ –ø–æ–ª–Ω—ã –∏–ª–∏ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã")
            raise ValueError()
    
    
    beta = np.arcsin(lambd*np.sin(phi))
    
    
    Pj = -Mvp*j
    try:
       Pg = 0.25*np.pi*D**2*(pg-100000)
    except:
       #tk.messagebox.showerror("–û—à–∏–±–∫–∞", "–ü–∞—Ä–∞–º–µ—Ç—Ä—ã –Ω–µ –ø–æ–ª–Ω—ã –∏–ª–∏ –Ω–µ–∫–æ—Ä—Ä–µ–∫—Ç–Ω—ã")
       raise ValueError() 
    
    M_rot_g = Pg*np.sin(phi+beta)*R/np.cos(beta)
    M_rot_j = Pj*np.sin(phi+beta)*R/np.cos(beta)

    #—Ä–∞–∑–ª–æ–∂–µ–Ω–∏–µ –≤ —Ä—è–¥ (–≥–∞–∑–æ–≤–∞—è —Å–æ—Å—Ç–∞–≤–ª—è—é—â–∞—è)
    Ag = np.zeros(K)
    Bg = np.zeros(K)
    for k in range(K):
        for i in range(N):
            Ag[k]+= 2/N*M_rot_g[i]*np.cos((k+1)*2*np.pi*(i+1)/N)
            Bg[k]+= 2/N*M_rot_g[i]*np.sin((k+1)*2*np.pi*(i+1)/N)

    M_rot_g_sqr = np.sqrt(Ag**2+Bg**2)
    phi_g = np.arctan(Ag/Bg)

    #—Å—Ä–µ–¥–Ω–∏–π –∫—Ä—É—Ç—è—â–∏–π –≥–∞–∑–æ–≤—ã–π –º–æ–º–µ–Ω—Ç
    M_rot_g_mean = (1/N)*sum(M_rot_g)

    #—Ä–∞–∑–ª–æ–∂–µ–Ω–∏–µ –≤ —Ä—è–¥ (–∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω–∞—è —Å–æ—Å—Ç–∞–≤–ª—è—é—â–∞—è)
    Aj = np.zeros(K)
    Bj = np.zeros(K)
    for k in range(K):
        for i in range(N):
            Aj[k]+= 2/N*M_rot_j[i]*np.cos((k+1)*2*np.pi*(i+1)/N)
            Bj[k]+= 2/N*M_rot_j[i]*np.sin((k+1)*2*np.pi*(i+1)/N)

    M_rot_j_sqr = np.sqrt(Aj**2+Bj**2)
    phi_j = np.arctan(Aj/Bj)

    #—Å—Ä–µ–¥–Ω–∏–π –∫—Ä—É—Ç—è—â–∏–π –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã–π –º–æ–º–µ–Ω—Ç
    M_rot_j_mean = 1/N*sum(M_rot_j)

    #–°—Ä–µ–¥–Ω–µ–µ –¥–∞–≤–ª–µ–Ω–∏–µ –≤ —Ä–∞–±–æ—á–µ–º –¥–∏–∞–ø–∞–∑–æ–Ω–µ
    n_diap=np.linspace(n_min, n_max, L)

    P_e=(30*tau/(V*I))*(Ne_nom/n_nom)*(a_coef+b_coef*n_diap/n_nom+c_coef*(n_diap/n_nom)**2)
    P_e_nom=(30*tau/(V*I))*(Ne_nom/n_nom)*(a_coef+b_coef+c_coef)

    #–≤—ã—á–∏—Å–ª—è–µ–º –≥–∞—Ä–º–æ–Ω–∏–∫–∏ –≥–∞–∑–æ–≤—ã—Ö —Å–æ—Å—Ç–∞–≤–ª—è—é—â–∏—Ö
    M_rot_n_g_strong_1=M_rot_g_sqr[K_m_strong_1-1]*P_e/P_e_nom
    M_rot_n_g_strong_2=M_rot_g_sqr[K_m_strong_2-1]*P_e/P_e_nom
    M_rot_n_g_strong_3=M_rot_g_sqr[K_m_strong_3-1]*P_e/P_e_nom

    M_rot_n_g_main_1=M_rot_g_sqr[K_m_main_1-1]*P_e/P_e_nom
    M_rot_n_g_main_2=M_rot_g_sqr[K_m_main_2-1]*P_e/P_e_nom
    M_rot_n_g_main_3=M_rot_g_sqr[K_m_main_3-1]*P_e/P_e_nom

    #–≤—ã—á–∏—Å–ª—è–µ–º –≥–∞—Ä–º–æ–Ω–∏–∫–∏ –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö —Å–æ—Å—Ç–∞–≤–ª—è—é—â–∏—Ö
    M_rot_n_j_strong_1=M_rot_j_sqr[round(K_m_strong_1)-1]*(n_diap/n_nom)**2
    M_rot_n_j_strong_2=M_rot_j_sqr[round(K_m_strong_2)-1]*(n_diap/n_nom)**2
    M_rot_n_j_strong_3=M_rot_j_sqr[round(K_m_strong_3)-1]*(n_diap/n_nom)**2

    M_rot_n_j_main_1=M_rot_j_sqr[round(K_m_main_1)-1]*(n_diap/n_nom)**2
    M_rot_n_j_main_2=M_rot_j_sqr[round(K_m_main_2)-1]*(n_diap/n_nom)**2
    M_rot_n_j_main_3=M_rot_j_sqr[round(K_m_main_3)-1]*(n_diap/n_nom)**2

    #7. –ê–º–ø–ª–∏—Ç—É–¥–∞ –∫–æ–ª–µ–±–∞–Ω–∏–π –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å

    A1_strong_1=(M_rot_n_j_strong_1+M_rot_n_g_strong_1)*sum(a)/(eps*K_m_strong_1*(2*np.pi*n_res_strong_1/60)*sum(a**2))
    A1_main_1=(M_rot_n_j_main_1+M_rot_n_g_main_1)*sum(a)/(eps*K_m_main_1*(2*np.pi*n_res_main_1/60)*sum(a**2))

    A1_strong_2=(M_rot_n_j_strong_2+M_rot_n_g_strong_2)*sum(a)/(eps*K_m_strong_2*(2*np.pi*n_res_strong_2/60)*sum(a**2))
    A1_main_2=(M_rot_n_j_main_2+M_rot_n_g_main_2)*sum(a)/(eps*K_m_main_2*(2*np.pi*n_res_main_2/60)*sum(a**2))

    A1_strong_3=(M_rot_n_j_strong_3+M_rot_n_g_strong_3)*sum(a)/(eps*K_m_strong_3*(2*np.pi*n_res_strong_3/60)*sum(a**2))
    A1_main_3=(M_rot_n_j_main_3+M_rot_n_g_main_3)*sum(a)/(eps*K_m_main_3*(2*np.pi*n_res_main_3/60)*sum(a**2))

    A_strong_1=np.zeros((M, L))
    A_main_1=np.zeros((M, L))
    A_strong_2=np.zeros((M, L))
    A_main_2=np.zeros((M, L))
    A_strong_3=np.zeros((M, L))
    A_main_3=np.zeros((M, L))

    for i in range(M):
        A_strong_1[i]=a[i]*A1_strong_1
        A_strong_2[i]=a[i]*A1_strong_2
        A_strong_3[i]=a[i]*A1_strong_3
        A_main_1[i]=a[i]*A1_main_1
        A_main_2[i]=a[i]*A1_main_2
        A_main_3[i]=a[i]*A1_main_3

    #—É–ø—Ä—É–≥–∏–µ –º–æ–º–µ–Ω—Ç—ã
    M_upr_strong_1=np.zeros((M-1, L))
    M_upr_main_1=np.zeros((M-1, L))
    M_upr_strong_2=np.zeros((M-1, L))
    M_upr_main_2=np.zeros((M-1, L))
    M_upr_strong_3=np.zeros((M-1, L))
    M_upr_main_3=np.zeros((M-1, L))

    for i in range(M-1):
        for j in range(L):
            M_upr_strong_1[i][j]=C[i]*(A_strong_1[i+1][j]-A_strong_1[i][j])
            M_upr_main_1[i][j]=C[i]*(A_main_1[i+1][j]-A_main_1[i][j])
            M_upr_strong_2[i][j]=C[i]*(A_strong_2[i+1][j]-A_strong_2[i][j])
            M_upr_main_2[i][j]=C[i]*(A_main_2[i+1][j]-A_main_2[i][j])
            M_upr_strong_3[i][j]=C[i]*(A_strong_3[i+1][j]-A_strong_3[i][j])
            M_upr_main_3[i][j]=C[i]*(A_main_3[i+1][j]-A_main_3[i][j])

    #–∫–∞—Å–∞—Ç–µ–ª—å–Ω—ã–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è
    T_upr_strong_1 = np.amax(np.abs(M_upr_strong_1))*16/np.pi/dk**3
    T_upr_main_1 = np.amax(np.abs(M_upr_main_1))*16/np.pi/dk**3
    T_upr_strong_2 = np.amax(np.abs(M_upr_strong_2))*16/np.pi/dk**3
    T_upr_main_2 = np.amax(np.abs(M_upr_main_2))*16/np.pi/dk**3
    T_upr_strong_3 = np.amax(np.abs(M_upr_strong_3))*16/np.pi/dk**3
    T_upr_main_3 = np.amax(np.abs(M_upr_main_3))*16/np.pi/dk**3

    #–º–∞–∫—Å–∏–º–∞–ª—å–Ω–æ–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏–µ
    T=np.amax(np.array([T_upr_strong_1, T_upr_main_1, T_upr_strong_2, T_upr_main_2, T_upr_strong_3, T_upr_main_3]))

    #9. –ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏ –∏ –ø–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞

    N_strong_1=(M_rot_g_sqr[round(K_m_strong_1)-1]+M_rot_j_sqr[round(K_m_strong_1)-1])*sum(a)*A1_strong_1*K_m_strong_1*w_s/2
    N_main_1=(M_rot_g_sqr[round(K_m_main_1)-1]+M_rot_j_sqr[round(K_m_main_1)-1])*sum(a)*A1_main_1*K_m_main_1*w_s/2

    N_strong_2=(M_rot_g_sqr[round(K_m_strong_2)-1]+M_rot_j_sqr[round(K_m_strong_2)-1])*sum(a)*A1_strong_2*K_m_strong_2*w_s/2
    N_main_2=(M_rot_g_sqr[round(K_m_main_2)-1]+M_rot_j_sqr[round(K_m_main_2)-1])*sum(a)*A1_main_2*K_m_main_2*w_s/2

    N_strong_3=(M_rot_g_sqr[round(K_m_strong_3)-1]+M_rot_j_sqr[round(K_m_strong_3)-1])*sum(a)*A1_strong_3*K_m_strong_3*w_s/2
    N_main_3=(M_rot_g_sqr[round(K_m_main_3)-1]+M_rot_j_sqr[round(K_m_main_3)-1])*sum(a)*A1_main_1*K_m_main_3*w_s/2

    N_sum=N_strong_1+N_main_1+N_strong_2+N_main_2+N_strong_3+N_main_3

    #–ø–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥

    Ne=Ne_nom*(a_coef+b_coef*n_diap/n_nom+c_coef*(n_diap/n_nom)**2)
    
    dg_strong_1=ge_nom*N_strong_1/Ne
    dg_main_1=ge_nom*N_main_1/Ne

    dg_strong_2=ge_nom*N_strong_2/Ne
    dg_main_2=ge_nom*N_main_2/Ne

    dg_strong_3=ge_nom*N_strong_3/Ne
    dg_main_3=ge_nom*N_main_3/Ne
    
    dg_sum=dg_strong_1+dg_main_1+dg_strong_2+dg_main_2+dg_strong_3+dg_main_3

    if np.isnan(w_s):
            raise ValueError()

    if textfile_on.get() or excelfile_on.get():
    
            # –°–æ–∑–¥–∞–Ω–∏–µ –Ω–∞–∑–≤–∞–Ω–∏—è –ø–∞–ø–∫–∏
            folder_name = datetime.now().strftime("%m-%d-%H-%M-%S")

           # –°–æ–∑–¥–∞–Ω–∏–µ –Ω–æ–≤–æ–π –ø–∞–ø–∫–∏
            os.mkdir(folder_name)

           # –ü–µ—Ä–µ—Ö–æ–¥ –≤ –Ω–æ–≤—É—é –ø–∞–ø–∫—É
            os.chdir(folder_name)

            if textfile_on.get():
                open(f_tables, 'w')
                open(f_out, 'w')
       
    print_Jp()
    print_Ji()
    print_JlC()
    print_w_s()
    print_a()
    print_n_res()
    print_M_rot_g_mean()
    print_M_rot_j_mean()
    print_Fourier_table()
    print_M_rot_n_g()
    print_M_rot_n_j()
    for i in range(M):
        print_Ai(i)
    if excelfile_on.get():
        print_excel_Ai()
        print_excel_M_upr()
    for i in range(M-1):
        print_M_upr(i)
    print_T()
    print_N()
    print_dg()
    if textfile_on.get() or excelfile_on.get():
            os.chdir('..')
    params_to_path(params_file_default_name)
    graphics()

    #plt.figure(num='A1_main_1')
    #plt.plot(n_diap, A1_main_1)
    #plt.title('A1_main_1')
    #plt.show()

#–û–ø—Ü–∏—è –ø–æ—Å—Ç—Ä–æ–µ–Ω–∏—è –≥—Ä–∞—Ñ–∏–∫–æ–≤
def graphics():
    global graph_label, harm_label, masses_label, upr_label
    global graph_box, harm_box, masses_box, upr_box
    global graph, harm, masses, inter

    harm=['–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è']
    harm.append('–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è')
    harm.append('–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è')
    harm.append('–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è')
    harm.append('–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è')
    harm.append('–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è')

    graph_label = ttk.Label(root, text='–ü–æ—Å—Ç—Ä–æ–µ–Ω–∏–µ –≥—Ä–∞—Ñ–∏–∫–∞:')
    graph_button = ttk.Button(root, text='–ü–æ—Å—Ç—Ä–æ–∏—Ç—å', command=make_graph)

    harm_label = ttk.Label(root, text='–ì–∞—Ä–º–æ–Ω–∏–∫–∞:')
    harm_box = ttk.Combobox(root, width=40, values=harm)
    harm_box.set('–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è')

    masses=[]
    for i in range(M):
        masses.append(str(i+1))
    masses_label = ttk.Label(root, text='–ù–æ–º–µ—Ä –º–∞—Å—Å—ã:')
    masses_box = ttk.Combobox(root, width=40, values=masses)
    masses_box.set('1')

    inter=[]
    for i in range(M-1):
        inter.append(str(i+1))
    upr_label = ttk.Label(root, text='–ò–Ω—Ç–µ—Ä–≤–∞–ª:')
    upr_box = ttk.Combobox(root, width=40, values=inter)
    upr_box.set('1')

    graph = ['–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –≥–∞–∑–æ–≤–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞']
    graph.append('–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞')
    graph.append('–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫')
    graph.append('–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫')
    graph.append('–ê–º–ø–ª–∏—Ç—É–¥—ã –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å')
    graph.append('–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ —É–ø—Ä—É–≥–∏–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è')
    graph.append('–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏')
    graph.append('–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞')
    graph.append('–°—É–º–º–∞—Ä–Ω—ã–µ –ø–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏')
    graph.append('–°—É–º–º–∞—Ä–Ω—ã–π –ø–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞')
    
    graph_box = ttk.Combobox(root, width=40, values=graph)
    graph_box.bind("<<ComboboxSelected>>", graphboxchanged)
    
    graph_label.grid(column=4, row=7, sticky=tk.W, padx=10)
    
    graph_box.set('–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –≥–∞–∑–æ–≤–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞')
    graph_box.grid(column=4, row=8, sticky=tk.W, padx=10, pady=5)
    graph_box.state(["readonly"])
    harm_box.state(["readonly"])
    masses_box.state(["readonly"])

    graph_button.grid(column=4, row=15, columnspan=1, pady=10)
        
def graphboxchanged(event):
    global graph_label, harm_label, masses_label, upr_label
    global graph_box, harm_box, masses_box, upr_box
    global graph, harm, masses, inter
    
    list1=['–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –≥–∞–∑–æ–≤–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞', '–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞',
           '–°—É–º–º–∞—Ä–Ω—ã–µ –ø–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏', '–°—É–º–º–∞—Ä–Ω—ã–π –ø–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞']
    list2=['–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫', '–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫',
           '–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏', '–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞']
    list3=['–ê–º–ø–ª–∏—Ç—É–¥—ã –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å']
    list4=['–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ —É–ø—Ä—É–≥–∏–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è']
    
    if graph_box.get() in list1:
        harm_label.grid_forget()
        harm_box.grid_forget()
        masses_label.grid_forget()
        masses_box.grid_forget()
        upr_label.grid_forget()
        upr_box.grid_forget()
        
    if graph_box.get() in list2:
        harm_label.grid(column=4, row=9, sticky=tk.W, padx=10)
        harm_box.grid(column=4, row=10, sticky=tk.W, padx=10, pady=5)
        masses_label.grid_forget()
        masses_box.grid_forget()
        upr_label.grid_forget()
        upr_box.grid_forget()
        
    if graph_box.get() in list3:
        upr_label.grid_forget()
        upr_box.grid_forget()
        harm_label.grid(column=4, row=9, sticky=tk.W, padx=10)
        harm_box.grid(column=4, row=10, sticky=tk.W, padx=10, pady=5)
        masses_label.grid(column=4, row=11, sticky=tk.W, padx=10)
        masses_box.grid(column=4, row=12, sticky=tk.W, padx=10, pady=5)
        
    if graph_box.get() in list4:
        masses_label.grid_forget()
        masses_box.grid_forget()
        harm_label.grid(column=4, row=9, sticky=tk.W, padx=10)
        harm_box.grid(column=4, row=10, sticky=tk.W, padx=10, pady=5)
        upr_label.grid(column=4, row=11, sticky=tk.W, padx=10)
        upr_box.grid(column=4, row=12, sticky=tk.W, padx=10, pady=5)
        
def make_graph():
    global graph_label, harm_label, masses_label, upr_label, dN_label
    global graph_box, harm_box, masses_box, upr_box
    global graph, harm, masses, inter

    if graph_box.get()=='–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –≥–∞–∑–æ–≤–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞':
        x = list(range(1, K+1))
        plt.figure(num='–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –≥–∞–∑–æ–≤–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞')
        plt.grid(True)
        plt.plot(x, M_rot_g_sqr)
        plt.title('–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –≥–∞–∑–æ–≤–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞')
        plt.xlabel('–ù–æ–º–µ—Ä –≥–∞—Ä–º–æ–Ω–∏–∫–∏')
        plt.ylabel('–ú–æ–º–µ–Ω—Ç, –ù\u00B7–º')
        plt.show()
        
    if graph_box.get()=='–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞':
        x = list(range(1, 25))
        plt.figure(num='–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞')
        plt.grid(True)
        plt.plot(x, M_rot_j_sqr)
        plt.title('–ì–∞—Ä–º–æ–Ω–∏–∫–∏ –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω–æ–≥–æ –º–æ–º–µ–Ω—Ç–∞')
        plt.xlabel('–ù–æ–º–µ—Ä –≥–∞—Ä–º–æ–Ω–∏–∫–∏')
        plt.ylabel('–ú–æ–º–µ–Ω—Ç, –ù\u00B7–º')
        plt.show()

    dict_Mng = {'–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è':M_rot_n_g_strong_1,
        '–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è':M_rot_n_g_main_1,
        '–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è':M_rot_n_g_strong_2,
        '–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è':M_rot_n_g_main_2,
        '–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è':M_rot_n_g_strong_3,
        '–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è':M_rot_n_g_main_3}
    
    if graph_box.get()=='–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫':
        x = n_diap
        plt.figure(num='–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Mng[harm_box.get()])
        plt.title('–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –≥–∞–∑–æ–≤—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ú–æ–º–µ–Ω—Ç, –ù\u00B7–º')
        plt.show()

    dict_Mnj = {'–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è':M_rot_n_j_strong_1,
        '–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è':M_rot_n_j_main_1,
        '–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è':M_rot_n_j_strong_2,
        '–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è':M_rot_n_j_main_2,
        '–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è':M_rot_n_j_strong_3,
        '–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è':M_rot_n_j_main_3}
    
    if graph_box.get()=='–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫':
        x = n_diap
        plt.figure(num='–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Mnj[harm_box.get()])
        plt.title('–ú–æ–º–µ–Ω—Ç—ã –≥–ª–∞–≤–Ω—ã—Ö –∏–Ω–µ—Ä—Ü–∏–æ–Ω–Ω—ã—Ö –≥–∞—Ä–º–æ–Ω–∏–∫')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ú–æ–º–µ–Ω—Ç, –ù\u00B7–º')
        plt.show()

    i=int(masses_box.get())-1
    dict_Ai = {'–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è':A_strong_1[i],
        '–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è':A_main_1[i],
        '–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è':A_strong_2[i],
        '–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è':A_main_2[i],
        '–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è':A_strong_3[i],
        '–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è':A_main_3[i]
    }
    
    if graph_box.get()=='–ê–º–ø–ª–∏—Ç—É–¥—ã –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å':
        x = n_diap
        plt.figure(num='–ê–º–ø–ª–∏—Ç—É–¥—ã –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å. –ú–∞—Å—Å–∞ ' + str(masses_box.get())+'. '+harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Ai[harm_box.get()])
        plt.title('–ê–º–ø–ª–∏—Ç—É–¥—ã –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å. –ú–∞—Å—Å–∞ ' + str(masses_box.get())+'. '+harm_box.get()+'.')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ê–º–ø–ª–∏—Ç—É–¥–∞, –º')
        plt.show()

    i=int(upr_box.get())-1
    
    dict_Mi = {'–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è':M_upr_strong_1[i],
        '–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è':M_upr_main_1[i],
        '–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è':M_upr_strong_2[i],
        '–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è':M_upr_main_2[i],
        '–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è':M_upr_strong_3[i],
        '–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è':M_upr_main_3[i]
    }
    
    if graph_box.get()=='–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ —É–ø—Ä—É–≥–∏–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è':
        x = n_diap
        plt.figure(num='–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ —É–ø—Ä—É–≥–∏–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è. –ò–Ω—Ç–µ—Ä–≤–∞–ª ' + str(upr_box.get())+'. '+harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_Mi[harm_box.get()])
        plt.title('–î–æ–ø–æ–ª–Ω–∏—Ç–µ–ª—å–Ω—ã–µ —É–ø—Ä—É–≥–∏–µ –Ω–∞–ø—Ä—è–∂–µ–Ω–∏—è. –ò–Ω—Ç–µ—Ä–≤–∞–ª ' + str(upr_box.get())+'. '+harm_box.get()+'.')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ù–∞–ø—Ä—è–∂–µ–Ω–∏–µ, –ü–∞')
        plt.show()

    dict_dN = {'–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è':N_strong_1,
        '–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è':N_main_1,
        '–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è':N_strong_2,
        '–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è':N_main_2,
        '–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è':N_strong_3,
        '–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è':N_main_3}

    if graph_box.get()=='–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏':
        x = n_diap
        plt.figure(num='–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_dN[harm_box.get()])
        plt.title('–ü–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏. ' + harm_box.get()+'.')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ü–æ—Ç–µ—Ä—è –º–æ—â–Ω–æ—Å—Ç–∏, –í—Ç')
        plt.show()

    dict_dg = {'–ü–µ—Ä–≤–∞—è —Å–∏–ª—å–Ω–∞—è':dg_strong_1,
        '–ü–µ—Ä–≤–∞—è –≥–ª–∞–≤–Ω–∞—è':dg_main_1,
        '–í—Ç–æ—Ä–∞—è —Å–∏–ª—å–Ω–∞—è':dg_strong_2,
        '–í—Ç–æ—Ä–∞—è –≥–ª–∞–≤–Ω–∞—è':dg_main_2,
        '–¢—Ä–µ—Ç—å—è —Å–∏–ª—å–Ω–∞—è':dg_strong_3,
        '–¢—Ä–µ—Ç—å—è –≥–ª–∞–≤–Ω–∞—è':dg_main_3}

    if graph_box.get()=='–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞':
        x = n_diap
        plt.figure(num='–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞. ' + harm_box.get()+'.')
        plt.grid(True)
        plt.plot(x, dict_dg[harm_box.get()])
        plt.title('–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞. ' + harm_box.get()+'.')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥, –≥/(–∫–í—Ç\u00B7—á)')
        plt.show()

    if graph_box.get()=='–°—É–º–º–∞—Ä–Ω—ã–µ –ø–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏':
        x = n_diap
        plt.figure(num='–°—É–º–º–∞—Ä–Ω—ã–µ –ø–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏.')
        plt.grid(True)
        plt.plot(x, N_sum)
        plt.title('–°—É–º–º–∞—Ä–Ω—ã–µ –ø–æ—Ç–µ—Ä–∏ –º–æ—â–Ω–æ—Å—Ç–∏.')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ü–æ—Ç–µ—Ä—è –º–æ—â–Ω–æ—Å—Ç–∏, –í—Ç')
        plt.show()

    if graph_box.get()=='–°—É–º–º–∞—Ä–Ω—ã–π –ø–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞':
        x = n_diap
        plt.figure(num='–°—É–º–º–∞—Ä–Ω—ã–π –ø–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞.')
        plt.grid(True)
        plt.plot(x, dg_sum)
        plt.title('–°—É–º–º–∞—Ä–Ω—ã–π –ø–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞.')
        plt.xlabel('n, –æ–±/–º')
        plt.ylabel('–ü–µ—Ä–µ—Ä–∞—Å—Ö–æ–¥, –≥/(–∫–í—Ç\u00B7—á)')
        plt.show()
    
        
# –°–æ–∑–¥–∞–Ω–∏–µ –≥–ª–∞–≤–Ω–æ–≥–æ –æ–∫–Ω–∞
root = tk.Tk()
root.iconbitmap("dvs.ico")
root.title('–†–∞—Å—á–µ—Ç –î–í–°')
root. resizable(False, False)

check_number = (root.register(is_valid_number), "%P")
check_float = (root.register(is_valid_float), "%P")

# –°–æ–∑–¥–∞–Ω–∏–µ —Ç–µ–∫—Å—Ç–æ–≤—ã—Ö –º–µ—Ç–æ–∫ –∏ –ø–æ–ª–µ–π –≤–≤–æ–¥–∞
label_M = ttk.Label(root, text='–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ –º–æ—Ç–æ—Ä–Ω—ã—Ö –º–∞—Å—Å M =')
label_M.grid(column=0, row=0, sticky=tk.W, padx=10, pady=5)
entry_M = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_M.insert(0, M)
entry_M.grid(column=1, row=0, padx=10, pady=5)

label_G = ttk.Label(root, text='–ú–æ–¥—É–ª—å —Å–¥–≤–∏–≥–∞ G (–ì–ü–∞) =')
label_G.grid(column=2, row=0, sticky=tk.W)
entry_G = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_G.insert(0, G/10**9)
entry_G.grid(column=3, row=0, padx=10)

label_I = ttk.Label(root, text='–ö–æ–ª–∏—á–µ—Å—Ç–≤–æ —Ü–∏–ª–∏–Ω–¥—Ä–æ–≤ I =')
label_I.grid(column=0, row=1, sticky=tk.W, padx=10, pady=5)
entry_I = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_I.insert(0, I)
entry_I.grid(column=1, row=1, padx=10, pady=5)

label_dk = ttk.Label(root, text='–î–∏–∞–º–µ—Ç—Ä –∫–æ—Ä–µ–Ω–Ω–æ–π —à–µ–π–∫–∏ d (–º) =')
label_dk.grid(column=2, row=1, sticky=tk.W, pady=5)
entry_dk = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_dk.insert(0, dk)
entry_dk.grid(column=3, row=1, padx=10, pady=5)

label_JM = ttk.Label(root, text='–ú–æ–º–µ–Ω—Ç –∏–Ω–µ—Ä—Ü–∏–∏ –º–∞—Ö–æ–≤–∏–∫–∞ –¥–µ–º–ø—Ñ–µ—Ä–∞ J_M (–∫–≥\u00B7–º\u00B2) =')
label_JM.grid(column=0, row=2, sticky=tk.W, padx=10, pady=5)
entry_JM = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_JM.insert(0, J_M)
entry_JM.grid(column=1, row=2, padx=10, pady=5)

label_R = ttk.Label(root, text='–†–∞–¥–∏—É—Å –∫—Ä–∏–≤–æ—à–∏–ø–∞ R (–º) =')
label_R.grid(column=2, row=2, sticky=tk.W, pady=5)
entry_R = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_R.insert(0, R)
entry_R.grid(column=3, row=2, padx=10, pady=5)

label_D = ttk.Label(root, text='–î–∏–∞–º–µ—Ç—Ä —Ü–∏–ª–∏–Ω–¥—Ä–∞ D (–º) =')
label_D.grid(column=0, row=3, sticky=tk.W, padx=10, pady=5)
entry_D = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_D.insert(0, D)
entry_D.grid(column=1, row=3, padx=10, pady=5)

label_lambd = ttk.Label(root, text='–ü–∞—Ä–∞–º–µ—Ç—Ä –ø–æ–¥–æ–±–∏—è \u03BB =')
label_lambd.grid(column=2, row=3, sticky=tk.W, pady=5)
entry_lambd = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_lambd.insert(0, lambd)
entry_lambd.grid(column=3, row=3, padx=10, pady=5)

label_Mvp = ttk.Label(root, text='–ú–∞—Å—Å–∞ –¥–≤–∏–∂—É—â–∏—Ö—Å—è —á–∞—Å—Ç–µ–π Mvp (–∫–≥) =')
label_Mvp.grid(column=0, row=4, sticky=tk.W, padx=10, pady=5)
entry_Mvp = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_Mvp.insert(0, Mvp)
entry_Mvp.grid(column=1, row=4, padx=10, pady=5)

label_n_nom = ttk.Label(root, text='–ù–æ–º–∏–Ω–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è (–æ–±/–º) =')
label_n_nom.grid(column=2, row=4, sticky=tk.W, pady=5)
entry_n_nom = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_n_nom.insert(3, n_nom)
entry_n_nom.grid(column=3, row=4, padx=10, pady=5)

label_n_min = ttk.Label(root, text='–ú–∏–Ω–∏–º–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è (–æ–±/–º) =')
label_n_min.grid(column=0, row=5, sticky=tk.W, padx=10, pady=5)
entry_n_min = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_n_min.insert(0, n_min)
entry_n_min.grid(column=1, row=5, padx=10, pady=5)

label_n_max = ttk.Label(root, text='–ú–∞–∫—Å–∏–º–∞–ª—å–Ω–∞—è —á–∞—Å—Ç–æ—Ç–∞ –≤—Ä–∞—â–µ–Ω–∏—è (–æ–±/–º) =')
label_n_max.grid(column=2, row=5, sticky=tk.W, pady=5)
entry_n_max = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_n_max.insert(3, n_max)
entry_n_max.grid(column=3, row=5, padx=10, pady=5)

label_V = ttk.Label(root, text='–û–±—ä–µ–º —Ü–∏–ª–∏–Ω–¥—Ä–∞ V (–ª) =')
label_V.grid(column=0, row=6, sticky=tk.W, padx=10, pady=5)
entry_V = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_V.insert(0, V*1000)
entry_V.grid(column=1, row=6, padx=10, pady=5)

label_tau = ttk.Label(root, text='–¢–∞–∫—Ç–Ω–æ—Å—Ç—å \u03C4 =')
label_tau.grid(column=2, row=6, sticky=tk.W, pady=5)
entry_tau = ttk.Entry(root, validate="key", validatecommand=check_number)
entry_tau.insert(3, tau)
entry_tau.grid(column=3, row=6, padx=10, pady=5)

label_Ne_nom = ttk.Label(root, text='–ù–æ–º–∏–Ω–∞–ª—å–Ω–∞—è –º–æ—â–Ω–æ—Å—Ç—å Ne_nom (–∫–í—Ç) =')
label_Ne_nom.grid(column=0, row=7, sticky=tk.W, padx=10, pady=5)
entry_Ne_nom = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_Ne_nom.insert(0, Ne_nom/1000)
entry_Ne_nom.grid(column=1, row=7, padx=10, pady=5)

label_eps = ttk.Label(root, text='–ö–æ—ç—Ñ—Ñ–∏—Ü–∏–µ–Ω—Ç –¥–µ–º–ø—Ñ–∏—Ä–æ–≤–∞–Ω–∏—è \u03B5 =')
label_eps.grid(column=2, row=7, sticky=tk.W, pady=5)
entry_eps = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_eps.insert(0, eps)
entry_eps.grid(column=3, row=7, padx=10, pady=5)

label_ge_nom = ttk.Label(root, text='–ù–æ–º–∏–Ω–∞–ª—å–Ω—ã–π —Ä–∞—Å—Ö–æ–¥ —Ç–æ–ø–ª–∏–≤–∞ –≤ –≥/(–∫–í—Ç\u00B7—á) =')
label_ge_nom.grid(column=0, row=8, sticky=tk.W, padx=10, pady=5)
entry_ge_nom = ttk.Entry(root, validate="key", validatecommand=check_float)
entry_ge_nom.insert(0, ge_nom)
entry_ge_nom.grid(column=1, row=8, padx=10, pady=5)

label_ge_nom = ttk.Label(root, text='–¢–∏–ø –¥–≤–∏–≥–∞—Ç–µ–ª—è:')
label_ge_nom.grid(column=2, row=8, sticky=tk.W, pady=5)
engine=['–ò—Å–∫—Ä–æ–≤–æ–π', '–î–∏–∑–µ–ª—å–Ω—ã–π']
engine_box = ttk.Combobox(root, values=engine)
engine_box.grid(column=3, row=8, sticky=tk.W, padx=10, pady=5)
engine_box.set('–ò—Å–∫—Ä–æ–≤–æ–π')
engine_box.state(["readonly"])

label_ge_nom = ttk.Label(root, text='–ö–æ–º–ø–æ–Ω–æ–≤–æ—á–Ω–∞—è —Å—Ö–µ–º–∞:')
label_ge_nom.grid(column=2, row=9, sticky=tk.W, pady=5)
scheme=['R', 'V', 'O']
scheme_box = ttk.Combobox(root, values=scheme)
scheme_box.grid(column=3, row=9, sticky=tk.W, padx=10, pady=5)
scheme_box.set('R')
scheme_box.state(["readonly"])

label_file = ttk.Label(root, text='–í—ã–±–µ—Ä–∏—Ç–µ —Ç–∏–ø—ã —Ñ–∞–π–ª–æ–≤ –¥–ª—è –≤—ã–≤–æ–¥–∞:')
label_file.grid(column=0, row=15, sticky=tk.W, padx=10, pady=5)

textfile_on = tk.BooleanVar()
text_checkbutton = ttk.Checkbutton(root, text='–¢–µ–∫—Å—Ç–æ–≤—ã–π —Ñ–∞–π–ª', variable=textfile_on)
#text_checkbutton.pack()
text_checkbutton.grid(column=1, row=15, padx=10, pady=5)

excelfile_on = tk.BooleanVar()
excel_checkbutton = ttk.Checkbutton(root, text='–§–∞–π–ª Excel', variable=excelfile_on)
excel_checkbutton.grid(column=2, row=15, padx=10, pady=5)


# –°–æ–∑–¥–∞–Ω–∏–µ –∫–Ω–æ–ø–∫–∏ –¥–ª—è —Ä–∞—Å—á–µ—Ç–∞ —Å—É–º–º—ã
calculate_button = ttk.Button(root, text='–†–∞—Å—Å—á–∏—Ç–∞—Ç—å', command=try_calculate)
calculate_button.grid(column=3, row=15, columnspan=1, pady=10)

# –°–æ–∑–¥–∞–Ω–∏–µ —Ç–µ–∫—Å—Ç–æ–≤–æ–≥–æ –ø–æ–ª—è –¥–ª—è –æ—Ç–æ–±—Ä–∞–∂–µ–Ω–∏—è —Ä–µ–∑—É–ª—å—Ç–∞—Ç–∞
Jp_label = ttk.Label(root)
w_s_label = ttk.Label(root)
M_rot_g_mean_label = ttk.Label(root)
M_rot_j_mean_label = ttk.Label(root)
T_label = ttk.Label(root)


# –°–æ–∑–¥–∞–Ω–∏–µ –º–µ–Ω—é
menu_bar = Menu(root)
root.config(menu=menu_bar)

# –î–æ–±–∞–≤–ª–µ–Ω–∏–µ –≤—ã–ø–∞–¥–∞—é—â–µ–≥–æ –º–µ–Ω—é "–§–∞–π–ª"
file_menu = Menu(menu_bar, tearoff=0)
menu_bar.add_cascade(label="–§–∞–π–ª", menu=file_menu)


# –§—É–Ω–∫—Ü–∏—è –¥–ª—è –∑–∞–∫—Ä—ã—Ç–∏—è –≤—Å–µ—Ö –æ–∫–æ–Ω
def close_all_windows():
    plt.close('all')
    root.quit()
    root.destroy()
    sys.exit()

def show_help():
    # –õ–æ–≥–∏–∫–∞ –æ—Ç–æ–±—Ä–∞–∂–µ–Ω–∏—è —Å–ø—Ä–∞–≤–∫–∏
    help_window = tk.Toplevel(root)
    help_window.resizable(False, False)
    help_window.iconbitmap("dvs.ico")
    help_text = tk.Text(help_window, wrap="word")
    help_text.insert(tk.END, " 1. –°–Ω–∞—á–∞–ª–∞ —á–µ—Ä–µ–∑ –º–µ–Ω—é –§–∞–π–ª –∑–∞–≥—Ä—É–∂–∞—é—Ç—Å—è —Å–Ω–æ–≤–Ω—ã–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã (–ø—Ä–µ–¥—ã–¥—É—â–∏–µ –æ—Å–Ω–æ–≤–Ω—ã–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã \
—Å–∏—Å—Ç–µ–º—ã –ø–æ–¥–≥—Ä—É–∂–∞—é—Ç—Å—è –∞–≤—Ç–æ–º–∞—Ç–∏—á–µ—Å–∫–∏ –∏–∑ —Ñ–∞–π–ª–∞ params.xlsx)\n 2. –ü–æ–ª—å–∑–æ–≤–∞—Ç–µ–ª—å –º–æ–∂–µ—Ç –≤—ã–±—Ä–∞—Ç—å —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏–µ —Ä–µ–∑—É–ª—å—Ç–∞—Ç–æ–≤ —Ä–∞–±–æ—Ç—ã –ø—Ä–æ–≥—Ä–∞–º–º—ã –≤ —Ç–µ–∫—Å—Ç–æ–≤—ã–π (output.txt) –∏/–∏–ª–∏ Excel (output.xlsx) —Ñ–∞–π–ª, \
–≤ –ø—Ä–æ—Ç–∏–≤–Ω–æ–º —Å–ª—É—á–∞–µ –æ–Ω–∏ –±—É–¥—É—Ç –≤—ã–≤–æ–¥–∏—Ç—å—Å—è —Ç–æ–ª—å–∫–æ –≤ –∫–æ–Ω—Å–æ–ª–∏ –ø—Ä–æ–≥—Ä–∞–º–º—ã\n 3. –î–ª—è —Ä–∞—Å—Å—á–µ—Ç–∞ –Ω–µ–æ–±—Ö–æ–¥–∏–º–æ –≤—ã–±—Ä–∞—Ç—å —Ñ–∞–π–ª –¥–∞–Ω–Ω—ã—Ö, –Ω–∞–ø—Ä–∏–º–µ—Ä testdata.xlsx\n 4. –ü–æ—Å–ª–µ —É—Å–ø–µ—à–Ω–æ–≥–æ \
—Ä–∞—Å—Å—á–µ—Ç–∞ –æ—Å–Ω–æ–≤–Ω—ã–µ –ø–∞—Ä–∞–º–µ—Ç—Ä—ã —Å–æ—Ö—Ä–∞–Ω—è—é—Ç—Å—è –≤ params.xlsx –∏ –æ—Ç–∫—Ä—ã–≤–∞–µ—Ç—Å—è –æ–ø—Ü–∏—è –ø–æ—Å—Ç—Ä–æ–µ–Ω–∏—è –∏ —Å–æ—Ö—Ä–∞–Ω–µ–Ω–∏—è –≥—Ä–∞—Ñ–∏–∫–æ–≤"    )
    help_text.config(state='disabled')
    help_text.grid()

menu_bar = tk.Menu(root)
file_menu = tk.Menu(menu_bar, tearoff=0)
file_menu.add_command(label="–°–æ—Ö—Ä–∞–Ω–∏—Ç—å", command=save_parameters)
file_menu.add_command(label="–ó–∞–≥—Ä—É–∑–∏—Ç—å", command=load_parameters)
file_menu.add_separator()
file_menu.add_command(label="–í—ã—Ö–æ–¥", command=close_all_windows)

help_menu = tk.Menu(menu_bar, tearoff=0)
help_menu.add_command(label="–°–ø—Ä–∞–≤–∫–∞", command=show_help)

menu_bar.add_cascade(label="–§–∞–π–ª", menu=file_menu)
menu_bar.add_cascade(label="–°–ø—Ä–∞–≤–∫–∞", menu=help_menu)

root.config(menu=menu_bar)

load_params(params_file_default_name)

# –ó–∞–ø—É—Å–∫ –≥–ª–∞–≤–Ω–æ–≥–æ —Ü–∏–∫–ª–∞ –æ–±—Ä–∞–±–æ—Ç–∫–∏ —Å–æ–±—ã—Ç–∏–π
root.mainloop()
